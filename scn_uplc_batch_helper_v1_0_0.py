#! python3

# All code written by Richard Pearson.
# For questions or comments:
# call or text: (650) 669-1460
# email: rpearson@scnutr.com

# Import libraries
from PyPDF2 import PdfFileWriter, PdfFileReader, PdfFileMerger
from reportlab.pdfgen import canvas
from reportlab.pdfgen.canvas import Canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.fonts import tt2ps
import pytz
import time
from time import sleep
from datetime import date, datetime, timedelta
from dateutil.relativedelta import *
import re
import os
from os import path
from tkinter import *
from tkinter.ttk import *
from tkinter import ttk
from tkinter import filedialog
from PIL import Image, ImageTk
from contextlib import suppress
import random
import pandas as pd
from ttkthemes import ThemedTk
import yaml
import msoffcrypto # pip install msoffcrypto-tool
import io
import openpyxl
import random
import pickle
import colorama
from colorama import Fore, Back, Style

def write_dict_to_yaml(dct, filepath):
	with open(filepath, 'w') as file:
		yaml.dump(dct, file, default_flow_style=False)

def read_yaml_to_dict(filepath):
	with open(filepath, 'r') as file:
		return yaml.safe_load(file)

def write_list_to_yaml(lst, filepath):
	with open(filepath, 'w') as file:
		yaml.dump(dict(lst), file, default_flow_style=False)

def read_yaml_to_list(filepath):
	with open(filepath, 'r') as file:
		return list(yaml.safe_load(file).items())

colorama.init(autoreset=False)

global R_drive_parent_folder
R_drive_parent_folder = os.path.normpath('R:\\QC\\Laboratory\\7. Lab improvement\\Raw Material Testing Plan\\test\\')

if os.path.exists(R_drive_parent_folder) == False:
	os.makedirs(R_drive_parent_folder, exist_ok=True)
	print(f'{Fore.GREEN}Directory Created:{Fore.RESET} {R_drive_parent_folder}\n')

primary_path = os.path.join(R_drive_parent_folder, 'config')
backup_path = os.path.join(os.getcwd(), 'config')
def yaml_data_xfer(filename='test.yaml', primary_path=primary_path, backup_path=backup_path, data='', silent=False):
	'''filename: should be the name of the yaml file (string)
	primary_file_path: primary location for the yaml file (R:\)
	backup_file_path: optional path location for backup yaml file location (C:\)
	data: The data that you want to store or read in the yaml file.'''

	primary_yaml_file = os.path.join(primary_path, filename)
	backup_yaml_file = os.path.join(backup_path, filename)

	def check_dtype(file, d, do='read'):
		# Checks the type() of the data being read or written then uses logic to select the correct yaml function.

		# file: the file path and name of file you want to read or write.
		# d: the data being transfered (accepts dicts, lists)
		# do: 'read' or 'write'

		x = ''

		dtype = type(d)
		if dtype == dict:
			if do == 'read':
				x = read_yaml_to_dict(file)
			if do == 'write':
				write_dict_to_yaml(d, file)

		if dtype == list:
			if do == 'read':
				x = read_yaml_to_list(file)
			if do == 'write':
				write_list_to_yaml(d, file)

		if x != '':
			return x

	# Check if yaml file exists in the primary directory location
	if os.path.exists(primary_yaml_file) == True:
		# Read primary locations YAML file
		x = check_dtype(primary_yaml_file, data, do='read')
	# Check if yaml file exists in the backup directory location
	elif os.path.exists(backup_yaml_file) == True:
		if backup_path != '':
			if silent == False:
				print(f"{Fore.RED}Could not find {filename} file in {primary_path}.{Fore.RESET} Trying {backup_path} location now.")
			# Read primary locations YAML file
			x = check_dtype(backup_yaml_file, data, do='read')
	else:
		x = data

		if backup_path == '':
			if silent == False:
				print(f"{Fore.RED}Could not find {filename} file in {primary_path}. Defaulting to hard coded version.{Fore.RESET}")

		if backup_path != '':
			if silent == False:
				print(f"{Fore.RED}Could not find {filename} file in {primary_path} or {backup_path}. Defaulting to hard coded version.{Fore.RESET}")
			if os.path.exists(backup_path) == False:
				os.makedirs(backup_path, exist_ok=True)

			if silent == False:
				print(f"Writing {filename} file to {backup_path}. You can now find this file and edit in this directory.")
			check_dtype(backup_yaml_file, data, do='write')

		if os.path.exists(primary_path) == False:
			os.makedirs(primary_path, exist_ok=True)

		if silent == False:
			print(f"Writing {filename} file to {primary_path}. You can find this file and edit in the {primary_path} directory.")
		check_dtype(primary_yaml_file, data, do='write')

	# return the data
	return x

equipment_config_default_data_input = {'names':["", "Michael Groden", "Emily Whitaker", "Richard Pearson", "Gerry Reyes", "Jennie Flores", "Joseph Porfido", "Curtis Halpin"],
										'scales': ["", "002442", "003218", "002442 and 003218"],
										'uplc': ["", "UPLC 1", "UPLC 2", "UPLC 3", "UPLC 4", "UPLC 1 and 4", "UPLC 1 and 2", "UPLC 1,2,4", "UPLC 2 and 4"],
										'vitamin_list': ["", "Vitamin C (QAL1209)", "B1 (QAL0930)", "B2 (QAL0930)", "B3 (QAL0930)", "B5 (QAL0930)", "B6 (QAL0930)", "B6 (QAL0901)",
														"B7 (QAL0933)", "B9 (QAL0937)", "B9 (QAL1202)", "B12 (QAL1016)", "B12 (QAL1203)", "B12 (QAL1022)", "Caffeine (QAL1201)",
														"Chlor Acids (QAL0940)", "CoQ10 (QAL1020)", "Curcumin (QAL0939)", "Melatonin (QAL1014)", "Melatonin (MTH-0035)" ]}

user_preferences_default_data_input = {'name_selection':0, 'scale_selection':0, 'uplc_selection':0, 'GMP_crossouts_font_color': 'Red', 'print_standard_notebook_info': False, 'print_injection_volumes': False}

standard_notebook_info = {
							'EXAMPLE':{'notebook_number':'RP-027', 'stock_page':'83', 'working_page':'92'},
							'MTH-0035':{'notebook_number':'', 'stock_page':'', 'working_page':''},
							'QAL0900':{'notebook_number':'', 'stock_page':'', 'working_page':''},
							'QAL0901':{'notebook_number':'', 'stock_page':'', 'working_page':''},
							'QAL0902':{'notebook_number':'', 'stock_page':'', 'working_page':''},
							'QAL0904':{'notebook_number':'', 'stock_page':'', 'working_page':''},
							'QAL0910':{'notebook_number':'', 'stock_page':'', 'working_page':''},
							'QAL0930':{'notebook_number':'', 'stock_page':'', 'working_page':''},
							'QAL0933':{'notebook_number':'', 'stock_page':'', 'working_page':''},
							'QAL0937':{'notebook_number':'', 'stock_page':'', 'working_page':''},
							'QAL0939':{'notebook_number':'', 'stock_page':'', 'working_page':''},
							'QAL0940':{'notebook_number':'', 'stock_page':'', 'working_page':''},
							'QAL1014':{'notebook_number':'', 'stock_page':'', 'working_page':''},
							'QAL1016':{'notebook_number':'', 'stock_page':'', 'working_page':''},
							'QAL1020':{'notebook_number':'', 'stock_page':'', 'working_page':''},
							'QAL1022':{'notebook_number':'', 'stock_page':'', 'working_page':''},
							'QAL1201':{'notebook_number':'', 'stock_page':'', 'working_page':''},
							'QAL1202':{'notebook_number':'', 'stock_page':'', 'working_page':''},
							'QAL1203':{'notebook_number':'', 'stock_page':'', 'working_page':''},
							'QAL1209':{'notebook_number':'', 'stock_page':'', 'working_page':''}
						}

equipment_config = yaml_data_xfer(filename='admin_config.yaml', primary_path=primary_path, backup_path='', data=equipment_config_default_data_input)
user_preferences_config = yaml_data_xfer(filename='user_preferences_config.yaml', primary_path=backup_path, backup_path='', data=user_preferences_default_data_input)
notebook_info_config = yaml_data_xfer(filename='notebook_info_config.yaml', primary_path=backup_path, backup_path='', data=standard_notebook_info)

def set_font_size(string, location):
	# input --> string
	# output --> integer
	font_size = 12
	string_length = len(string)

	if location == 'uplc_batch_wip_field':
		if string_length <= 10:
			font_size = 12
		elif (string_length > 10) and (string_length <= 15):
			font_size = 10
		else:
			font_size = 8

	if location == 'uplc_batch_lot_field':
		if string_length <= 10:
			font_size = 12
		elif (string_length > 10) and (string_length <= 15):
			font_size = 10
		else:
			font_size = 8

	if location == 'uplc_batch_description_field':
		if string_length <= 45:
			font_size = 12
		else:
			font_size = 10

	return font_size

# ------------------------------------ Tkinter Start --------------------------------------------- #

# Hard coded configuration
VERSION = 'Version 1.0.0'

# Make a tkinter window for the program GUI
root = ThemedTk(theme='vista')

# Allow for styling of the GUI
# See the following link for style of comboboxes etc:
# https://stackoverflow.com/questions/31545559/how-to-change-background-color-in-ttk-comboboxs-listview

# our_themes = ttk.Style().theme_names()
# our_themes2 = root.get_themes()
selected_themes = ['vista', 'xpnative']

style = ttk.Style(root)

# Banner Title
root.title(f"UPLC Sample Form Creator - {VERSION}")

# Sets window width and height and also opens at center of screen
app_width = 1500
app_height = 680
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
screen_x = (screen_width / 2) - (app_width / 2)
screen_y = (screen_height / 2) - (app_height / 2) - 50
root.geometry(f'{app_width}x{app_height}+{int(screen_x)}+{int(screen_y)}')

# Make a dropdown menu for the GUI
main_menu = Menu(root)
root.config(menu=main_menu)

# Makes the excape key bind to exiting the program
def exit(event):
	root.destroy()

# Escape Binding
root.bind("<Escape>", exit)

def form_maker_view():
	hide_all_frames()
	root.title("UPLC Sample Form Creator")
	lbl.configure(text="UPLC Form Maker")
	window.grid(column=0, row=1, columnspan=2, sticky=W)
	misc_info_frame.grid(column=0, row=2, sticky=W)
	form_button_frame.grid(column=1, row=2)

def hide_all_frames():
	window.grid_forget()
	misc_info_frame.grid_forget()
	form_button_frame.grid_forget()

def show_user_guide():
	# Open the PDF User Guide Document
	#os.startfile(r".\help\UPLC Form Generator User Guide.docx") # NOTE: The os.startfile will only work on Windows Machines #
	os.startfile(r".\help\User Manual.docx")

# Create a menu item
file_menu = Menu(main_menu, tearoff=False)
main_menu.add_cascade(label="File", menu=file_menu)
file_menu.add_command(label="Exit", command=root.quit)

view_menu = Menu(main_menu, tearoff=False)
main_menu.add_cascade(label="View", menu=view_menu)
view_menu.add_command(label="UPLC Form Maker", command=form_maker_view)

main_theme_menu = Menu(main_menu, tearoff=False)
main_menu.add_cascade(label="Preferences", menu=main_theme_menu)

theme_menu = Menu(main_menu, tearoff=False)
main_theme_menu.add_cascade(label="Themes", menu=theme_menu)

# Make some frames
window = Frame(root)
window.grid(column=0, row=1, columnspan=2, sticky=W)

def change_theme(theme):
	style.theme_use(theme)
	theme_bg_color = style.lookup("TCanvas", "background")
	root.config(bg=theme_bg_color)
	root.update()

# Theme Sub menus
for t in selected_themes:
	theme_menu.add_command(label=t, command=lambda t=t: change_theme(t))

help_menu = Menu(main_menu, tearoff=False)
main_menu.add_cascade(label="Help", menu=help_menu)
help_menu.add_command(label="User Guide", command=show_user_guide)

####################################################################################################################################################################################
# #####----------------------------------------------------------------------------------------------------------------------------------------------------------------------##### #
# #####--------------------------------------------------------- FORM MAKER GRAPHICAL USER INTERFACE ------------------------------------------------------------------------##### #
# #####----------------------------------------------------------------------------------------------------------------------------------------------------------------------##### #
####################################################################################################################################################################################

# Sets a background image for the GUI
global background_image
#background_image = ImageTk.PhotoImage(file=r"images\scn_tagline.png")
background_image = ImageTk.PhotoImage(file=r"images\SCNBestcoLogo_resized.png")

header_frame = Frame(root)
header_frame.grid(column=1, row=0, columnspan=10, pady=20) # made a pady=20

global background
background = Label(root, image=background_image)
background.grid(column=0, row=0, columnspan=1, rowspan=1, padx=50, sticky=W)

# Sets 32x32 .png file as icon
global icon
#icon = PhotoImage(file=r"images\scn_icon.png")
icon = PhotoImage(file=r"images\scn_bestco_icon_resized_icon.png")
root.iconphoto(False, icon)

# Headers for GUI
lbl = Label(header_frame, text="UPLC Form Maker", font=("Arial", 16))
lbl.grid(column=2, row=0, columnspan=4, pady=30, sticky=W)

wip_header = Label(window, text="Sample WIP\n   or Item:", font=("Arial Bold", 10))
wip_header.grid(column=0, row=0, padx=(50,5))

lot_header = Label(window, text="Lot:", font=("Arial Bold", 10))
lot_header.grid(column=1, row=0, padx=5)

description_header = Label(window, text="Sample Description", font=("Arial Bold", 10))
description_header.grid(column=2, columnspan=2, row=0, padx=5)

lab_reports_header = Label(window, text="Lab Report ID", font=("Arial Bold", 10))
lab_reports_header.grid(column=4, row=0, padx=5)

sample_type_header = Label(window, text="Sample Type", font=("Arial Bold", 10))
sample_type_header.grid(column=5, row=0, padx=5)

analyte_type_header = Label(window, text="Analyte Type", font=("Arial Bold", 10))
analyte_type_header.grid(column=6, row=0, padx=5)

weight_header = Label(window, text="Piece Wt (g) or \n Usage (mg)", font=("Arial Bold", 10))
weight_header.grid(column=7, row=0, padx=5)

units_per_spec_header = Label(window, text="Units per\n Spec", font=("Arial Bold", 10))
units_per_spec_header.grid(column=8, row=0, padx=5)

triplicate_header = Label(window, text="Reps", font=("Arial Bold", 10))
triplicate_header.grid(column=9, row=0, padx=5)

stability_header = Label(window, text="Stability", font=("Arial Bold", 10))
stability_header.grid(column=10, row=0, padx=5)

# Misc info frame
misc_info_frame = Frame(root)
misc_info_frame.grid(column=0, row=2, sticky=W)

name_header = Label(misc_info_frame, text="Analyst Name", font=("Arial Bold", 10))
name_header.grid(column=0, row=0, padx=(50,5), pady=(20,5))

scale_header = Label(misc_info_frame, text="Scale No.", font=("Arial Bold", 10))
scale_header.grid(column=1, row=0, padx=5, pady=(20,5), sticky=W)

uplc_system_header = Label(misc_info_frame, text="UPLC System", font=("Arial Bold", 10))
uplc_system_header.grid(column=2, row=0, padx=5, ipadx=10, pady=(20,5), sticky=W)

date_selector_header = Label(misc_info_frame, text="Date", font=("Arial Bold", 10))
date_selector_header.grid(column=3, row=0, padx=5, pady=(20,5))

name = Combobox(misc_info_frame)
#name['values']= (name_list)
name['values']= (equipment_config['names'])
#name.current(user_preferences_config['name_selection']) #set the selected item
name.current(user_preferences_config['name_selection'])
name.grid(column=0, row=1, padx=(50,2), pady=5, sticky=W)

scale = Combobox(misc_info_frame, width=22)
scale['values'] = (equipment_config['scales'])
scale.current(user_preferences_config['scale_selection']) #set the selected item
scale.grid(column=1, row=1, padx=2, pady=5)

uplc_system_id = Combobox(misc_info_frame, width=15)
#uplc_system_id['values']= ("", "UPLC 1", "UPLC 2", "UPLC 3", "UPLC 4", "UPLC 1 and 4", "UPLC 1 and 2", "UPLC 1,2,4", "UPLC 2 and 4")
uplc_system_id['values']= (equipment_config['uplc'])
uplc_system_id.current(user_preferences_config['uplc_selection']) #set the selected item
uplc_system_id.grid(column=2, row=1, padx=2, pady=5, sticky=W)

# Display todays date in GUI and allow for changes in the date (only future up to 5 days)
todays_date = date.today()
one_day_in_future = date.today() + timedelta(1)
two_day_in_future = date.today() + timedelta(2)
two_day_future_weekday = " (" + two_day_in_future.strftime('%A') + ')'
three_day_in_future = date.today() + timedelta(3)
three_day_future_weekday = " (" + three_day_in_future.strftime('%A') + ')'
four_day_in_future = date.today() + timedelta(4)
four_day_future_weekday = " (" + four_day_in_future.strftime('%A') + ')'
date_string = todays_date.strftime("%m/%d/%Y")
one_day_in_future_string = one_day_in_future.strftime("%m/%d/%Y")
two_day_in_future_string = two_day_in_future.strftime("%m/%d/%Y")
three_day_in_future_string = three_day_in_future.strftime("%m/%d/%Y")
four_day_in_future_string = four_day_in_future.strftime("%m/%d/%Y")

date_selector = Combobox(misc_info_frame, width=24)
date_selector['values']=("", date_string + " (Today)", one_day_in_future_string + " (Tomorrow)", two_day_in_future_string + two_day_future_weekday, three_day_in_future_string + three_day_future_weekday, four_day_in_future_string + four_day_future_weekday)
date_selector.current(1)
date_selector.grid(column=3, row=1, padx=2, pady=5, sticky=W)


####################################################################################################################################################################################
# #####----------------------------------------------------------------------------------------------------------------------------------------------------------------------##### #
# #####--------------------------------------------------------- Form Maker Fields in 'window' Frame ------------------------------------------------------------------------##### #
# #####----------------------------------------------------------------------------------------------------------------------------------------------------------------------##### #
####################################################################################################################################################################################

# *************** Fat Soluble Vitamin Drop Down Lists *************** #

# make vitamin A list
vitamin_A_list = ["", "A (QAL0904)"] # "A Palmitate ()", "Beta Carotene ()"

# make vitamin C list
vitamin_C_list = ["", "Vitamin C (QAL0910)"]

# make vitamin D list
vitamin_D_list = ["", "D (QAL0902)"] # "D2 ()"

# make vitamin E list
vitamin_E_list = ["", "E (QAL0900)"]

# ************** Water Soluble Vitamin Drop Down Lists ************** #

# make a water soluble vitamin list
vitamin_list = equipment_config['vitamin_list']

################### Column widths ###################
# Column 1
wip_col_width = 20
fs1_col_width = wip_col_width - 3
#----------------------------------------------------
# Column 2
lot_col_width = 20
fs2_col_width = lot_col_width - 3
#----------------------------------------------------
# Column 3
description_col_width = 40
fs3_col_width = int(description_col_width / 2) - 2
fs4_col_width = int(description_col_width / 2) - 6
#----------------------------------------------------
# Column 4
lab_report_col_width = 24
ws1_col_width = lab_report_col_width - 3
#----------------------------------------------------
# Column 5
samp_type_col_width = 24
ws2_col_width = samp_type_col_width - 0
#----------------------------------------------------
# Column 6
analyte_type_col_width = 24
ws3_col_width = analyte_type_col_width - 0
#----------------------------------------------------
# Column 7
piece_weight_col_width = 24
ws4_col_width = piece_weight_col_width - 3
#----------------------------------------------------
# Column 8
units_col_width = 24
ws5_col_width = units_col_width - 3
#----------------------------------------------------

# Form maker window frame row 1 content
wip1 = Entry(window, width=wip_col_width)
wip1.grid(column=0, row=1, padx=(50,2), pady=5)

lot1 = Entry(window, width=lot_col_width)
lot1.grid(column=1, row=1, padx=2, pady=5)

desc1 = Entry(window, width=description_col_width)
desc1.grid(column=2, columnspan=2, row=1, padx=2, pady=5)

lr1 = Entry(window, width=lab_report_col_width)
lr1.grid(column=4, row=1, padx=2, pady=5)

st1 = Combobox(window, width=samp_type_col_width)
st1['values'] = ("", "Finished Product", "Raw Blend", "Percent Active")
st1['state'] = 'readonly'
st1.current(0) #set the selected item
st1.grid(column=5, row=1, padx=2, pady=5)

polarity1 = Combobox(window, width=analyte_type_col_width)
polarity1['values'] = ("", "Fat Soluble", "Water Soluble", "Other")
polarity1['state'] = 'readonly'
polarity1.current(0) #set the selected item
polarity1.grid(column=6, row=1, padx=2, pady=5)

wt1 = Entry(window, width=piece_weight_col_width)
wt1.grid(column=7, row=1, padx=2, pady=5)

ups1 = Entry(window, width=units_col_width)
ups1.grid(column=8, row=1, padx=2, pady=5)

chk_tri1 = BooleanVar()
chk_tri1.set(False) #set check state
chk = Checkbutton(window, text='', var=chk_tri1)
chk.grid(column=9, row=1, padx=2, pady=5)

stability_box1 = BooleanVar()
stability_box1.set(False) #set check state
stab = Checkbutton(window, text='', var=stability_box1)
stab.grid(column=10, row=1, padx=2, pady=5)

# Row 3 content
wip2 = Entry(window, width=wip_col_width)
wip2.grid(column=0, row=3, padx=(50,2), pady=5)

lot2 = Entry(window, width=lot_col_width)
lot2.grid(column=1, row=3, padx=2, pady=5)

desc2 = Entry(window, width=description_col_width)
desc2.grid(column=2, columnspan=2, row=3, padx=2, pady=5)

lr2 = Entry(window, width=lab_report_col_width)
lr2.grid(column=4, row=3, padx=2, pady=5)

st2 = Combobox(window, width=samp_type_col_width)
st2['values']= ("", "Finished Product", "Raw Blend", "Percent Active")
st2['state'] = 'readonly'
st2.current(0) #set the selected item
st2.grid(column=5, row=3, padx=2, pady=5)

polarity2 = Combobox(window, width=analyte_type_col_width)
polarity2['values']= ("", "Fat Soluble", "Water Soluble", "Other")
polarity2['state'] = 'readonly'
polarity2.current(0) #set the selected item
polarity2.grid(column=6, row=3, padx=2, pady=5)

wt2 = Entry(window, width=piece_weight_col_width)
wt2.grid(column=7, row=3, padx=2, pady=5)

ups2 = Entry(window, width=units_col_width)
ups2.grid(column=8, row=3, padx=2, pady=5)

chk_tri2 = BooleanVar()
chk_tri2.set(False) #set check state
chk = Checkbutton(window, text='', var=chk_tri2)
chk.grid(column=9, row=3, padx=2, pady=5)

stability_box2 = BooleanVar()
stability_box2.set(False) #set check state
stab = Checkbutton(window, text='', var=stability_box2)
stab.grid(column=10, row=3, padx=2, pady=5)

# Row 5 content
wip3 = Entry(window, width=wip_col_width)
wip3.grid(column=0, row=5, padx=(50,2), pady=5)

lot3 = Entry(window, width=lot_col_width)
lot3.grid(column=1, row=5, padx=2, pady=5)

desc3 = Entry(window, width=description_col_width)
desc3.grid(column=2, columnspan=2, row=5, padx=2, pady=5)

lr3 = Entry(window, width=lab_report_col_width)
lr3.grid(column=4, row=5, padx=2, pady=5)

st3 = Combobox(window, width=samp_type_col_width)
st3['values']= ("", "Finished Product", "Raw Blend", "Percent Active")
st3['state'] = 'readonly'
st3.current(0) #set the selected item
st3.grid(column=5, row=5, padx=2, pady=5)

polarity3 = Combobox(window, width=analyte_type_col_width)
polarity3['values']= ("", "Fat Soluble", "Water Soluble", "Other")
polarity3['state'] = 'readonly'
polarity3.current(0) #set the selected item
polarity3.grid(column=6, row=5, padx=2, pady=5)

wt3 = Entry(window, width=piece_weight_col_width)
wt3.grid(column=7, row=5, padx=2, pady=5)

ups3 = Entry(window, width=units_col_width)
ups3.grid(column=8, row=5, padx=2, pady=5)

chk_tri3 = BooleanVar()
chk_tri3.set(False) #set check state
chk = Checkbutton(window, text='', var=chk_tri3)
chk.grid(column=9, row=5, padx=2, pady=5)

stability_box3 = BooleanVar()
stability_box3.set(False) #set check state
stab = Checkbutton(window, text='', var=stability_box3)
stab.grid(column=10, row=5, padx=2, pady=5)

# Row 7 content
wip4 = Entry(window, width=wip_col_width)
wip4.grid(column=0, row=7, padx=(50,2), pady=5)

lot4 = Entry(window, width=lot_col_width)
lot4.grid(column=1, row=7, padx=2, pady=5)

desc4 = Entry(window, width=description_col_width)
desc4.grid(column=2, columnspan=2, row=7, padx=2, pady=5)

lr4 = Entry(window, width=lab_report_col_width)
lr4.grid(column=4, row=7, padx=2, pady=5)

st4 = Combobox(window, width=samp_type_col_width)
st4['values']= ("", "Finished Product", "Raw Blend", "Percent Active")
st4['state'] = 'readonly'
st4.current(0) #set the selected item
st4.grid(column=5, row=7, padx=2, pady=5)

polarity4 = Combobox(window, width=analyte_type_col_width)
polarity4['values']= ("", "Fat Soluble", "Water Soluble", "Other")
polarity4['state'] = 'readonly'
polarity4.current(0) #set the selected item
polarity4.grid(column=6, row=7, padx=2, pady=5)

wt4 = Entry(window, width=piece_weight_col_width)
wt4.grid(column=7, row=7, padx=2, pady=5)

ups4 = Entry(window, width=units_col_width)
ups4.grid(column=8, row=7, padx=2, pady=5)

chk_tri4 = BooleanVar()
chk_tri4.set(False) #set check state
chk = Checkbutton(window, text='', var=chk_tri4)
chk.grid(column=9, row=7, padx=2, pady=5)

stability_box4 = BooleanVar()
stability_box4.set(False) #set check state
stab = Checkbutton(window, text='', var=stability_box4)
stab.grid(column=10, row=7, padx=2, pady=5)

# Row 9 content
wip5 = Entry(window, width=wip_col_width)
wip5.grid(column=0, row=9, padx=(50,2), pady=5)

lot5 = Entry(window, width=lot_col_width)
lot5.grid(column=1, row=9, padx=2, pady=5)

desc5 = Entry(window, width=description_col_width)
desc5.grid(column=2, columnspan=2, row=9, padx=2, pady=5)

lr5 = Entry(window, width=lab_report_col_width)
lr5.grid(column=4, row=9, padx=2, pady=5)

st5 = Combobox(window, width=samp_type_col_width)
st5['values']= ("", "Finished Product", "Raw Blend", "Percent Active")
st5['state'] = 'readonly'
st5.current(0) #set the selected item
st5.grid(column=5, row=9, padx=2, pady=5)

polarity5 = Combobox(window, width=analyte_type_col_width)
polarity5['values']= ("", "Fat Soluble", "Water Soluble", "Other")
polarity5['state'] = 'readonly'
polarity5.current(0) #set the selected item
polarity5.grid(column=6, row=9, padx=2, pady=5)

wt5 = Entry(window, width=piece_weight_col_width)
wt5.grid(column=7, row=9, padx=2, pady=5)

ups5 = Entry(window, width=units_col_width)
ups5.grid(column=8, row=9, padx=2, pady=5)

chk_tri5 = BooleanVar()
chk_tri5.set(False) #set check state
chk = Checkbutton(window, text='', var=chk_tri5)
chk.grid(column=9, row=9, padx=2, pady=5)

stability_box5 = BooleanVar()
stability_box5.set(False) #set check state
stab = Checkbutton(window, text='', var=stability_box5)
stab.grid(column=10, row=9, padx=2, pady=5)

# Row 11 content
wip6 = Entry(window, width=wip_col_width)
wip6.grid(column=0, row=11, padx=(50,2), pady=5)

lot6 = Entry(window, width=lot_col_width)
lot6.grid(column=1, row=11, padx=2, pady=5)

desc6 = Entry(window, width=description_col_width)
desc6.grid(column=2, columnspan=2, row=11, padx=2, pady=5)

lr6 = Entry(window, width=lab_report_col_width)
lr6.grid(column=4, row=11, padx=2, pady=5)

st6 = Combobox(window, width=samp_type_col_width)
st6['values'] = ("", "Finished Product", "Raw Blend", "Percent Active")
st6['state'] = 'readonly'
st6.current(0) #set the selected item
st6.grid(column=5, row=11, padx=2, pady=5)

polarity6 = Combobox(window, width=analyte_type_col_width)
polarity6['values'] = ("", "Fat Soluble", "Water Soluble", "Other")
polarity6['state'] = 'readonly'
polarity6.current(0) #set the selected item
polarity6.grid(column=6, row=11, padx=2, pady=5)

wt6 = Entry(window, width=piece_weight_col_width)
wt6.grid(column=7, row=11, padx=2, pady=5)

ups6 = Entry(window, width=units_col_width)
ups6.grid(column=8, row=11, padx=2, pady=5)

chk_tri6 = BooleanVar()
chk_tri6.set(False) #set check state
chk = Checkbutton(window, text='', var=chk_tri6)
chk.grid(column=9, row=11, padx=2, pady=5)

stability_box6 = BooleanVar()
stability_box6.set(False) #set check state
stab = Checkbutton(window, text='', var=stability_box6)
stab.grid(column=10, row=11, padx=2, pady=5)

# Row 13 content
wip7 = Entry(window, width=wip_col_width)
wip7.grid(column=0, row=13, padx=(50,2), pady=5)

lot7 = Entry(window, width=lot_col_width)
lot7.grid(column=1, row=13, padx=2, pady=5)

desc7 = Entry(window, width=description_col_width)
desc7.grid(column=2, columnspan=2, row=13, padx=2, pady=5)

lr7 = Entry(window, width=lab_report_col_width)
lr7.grid(column=4, row=13, padx=2, pady=5)

st7 = Combobox(window, width=samp_type_col_width)
st7['values'] = ("", "Finished Product", "Raw Blend", "Percent Active")
st7['state'] = 'readonly'
st7.current(0) #set the selected item
st7.grid(column=5, row=13, padx=2, pady=5)

polarity7 = Combobox(window, width=analyte_type_col_width)
polarity7['values'] = ("", "Fat Soluble", "Water Soluble", "Other")
polarity7['state'] = 'readonly'
polarity7.current(0) #set the selected item
polarity7.grid(column=6, row=13, padx=2, pady=5)

wt7 = Entry(window, width=piece_weight_col_width)
wt7.grid(column=7, row=13, padx=2, pady=5)

ups7 = Entry(window, width=units_col_width)
ups7.grid(column=8, row=13, padx=2, pady=5)

chk_tri7 = BooleanVar()
chk_tri7.set(False) #set check state
chk = Checkbutton(window, text='', var=chk_tri7)
chk.grid(column=9, row=13, padx=2, pady=5)

stability_box7 = BooleanVar()
stability_box7.set(False) #set check state
stab = Checkbutton(window, text='', var=stability_box7)
stab.grid(column=10, row=13, padx=2, pady=5)

#################### DROPDOWN MENUS ####################

# make vitamin drop down menus
r1_fs1 = Combobox(window, width=fs1_col_width)
r1_fs1['values']= (vitamin_A_list)
r1_fs1.current(0) #set the selected item
r1_fs1.grid(column=0, row=2, padx=(50,2))

r1_fs2 = Combobox(window, width=fs2_col_width)
r1_fs2['values']= (vitamin_C_list)
r1_fs2.current(0) #set the selected item
r1_fs2.grid(column=1, row=2)

r1_fs3 = Combobox(window, width=fs3_col_width)
r1_fs3['values']= (vitamin_E_list)
r1_fs3.current(0) #set the selected item
r1_fs3.grid(column=2, row=2)

r1_fs4 = Combobox(window, width=fs4_col_width)
r1_fs4['values']= (vitamin_D_list)
r1_fs4.current(0) #set the selected item
r1_fs4.grid(column=3, row=2)

r1_ws1 = Combobox(window, width=ws1_col_width)
r1_ws1['values']= (vitamin_list)
r1_ws1.current(0) #set the selected item
r1_ws1.grid(column=4, row=2)

r1_ws2 = Combobox(window, width=ws2_col_width)
r1_ws2['values']= (vitamin_list)
r1_ws2.current(0) #set the selected item
r1_ws2.grid(column=5, row=2)

r1_ws3 = Combobox(window, width=ws3_col_width)
r1_ws3['values']= (vitamin_list)
r1_ws3.current(0) #set the selected item
r1_ws3.grid(column=6, row=2)

r1_ws4 = Combobox(window, width=ws4_col_width)
r1_ws4['values']= (vitamin_list)
r1_ws4.current(0) #set the selected item
r1_ws4.grid(column=7, row=2)

r1_ws5 = Combobox(window, width=ws5_col_width)
r1_ws5['values']= (vitamin_list)
r1_ws5.current(0) #set the selected item
r1_ws5.grid(column=8, row=2)

# make the second set of drop downs
r2_fs1 = Combobox(window, width=fs1_col_width)
r2_fs1['values']= (vitamin_A_list)
r2_fs1.current(0) #set the selected item
r2_fs1.grid(column=0, row=4, padx=(50,2))

r2_fs2 = Combobox(window, width=fs2_col_width)
r2_fs2['values']= (vitamin_C_list)
r2_fs2.current(0) #set the selected item
r2_fs2.grid(column=1, row=4)

r2_fs3 = Combobox(window, width=fs3_col_width)
r2_fs3['values']= (vitamin_E_list)
r2_fs3.current(0) #set the selected item
r2_fs3.grid(column=2, row=4)

r2_fs4 = Combobox(window, width=fs4_col_width)
r2_fs4['values']= (vitamin_D_list)
r2_fs4.current(0) #set the selected item
r2_fs4.grid(column=3, row=4)

r2_ws1 = Combobox(window, width=ws1_col_width)
r2_ws1['values']= (vitamin_list)
r2_ws1.current(0) #set the selected item
r2_ws1.grid(column=4, row=4)

r2_ws2 = Combobox(window, width=ws2_col_width)
r2_ws2['values']= (vitamin_list)
r2_ws2.current(0) #set the selected item
r2_ws2.grid(column=5, row=4)

r2_ws3 = Combobox(window, width=ws3_col_width)
r2_ws3['values']= (vitamin_list)
r2_ws3.current(0) #set the selected item
r2_ws3.grid(column=6, row=4)

r2_ws4 = Combobox(window, width=ws4_col_width)
r2_ws4['values']= (vitamin_list)
r2_ws4.current(0) #set the selected item
r2_ws4.grid(column=7, row=4)

r2_ws5 = Combobox(window, width=ws5_col_width)
r2_ws5['values']= (vitamin_list)
r2_ws5.current(0) #set the selected item
r2_ws5.grid(column=8, row=4)

# make third set of dropdowns
r3_fs1 = Combobox(window, width=fs1_col_width)
r3_fs1['values']= (vitamin_A_list)
r3_fs1.current(0) #set the selected item
r3_fs1.grid(column=0, row=6, padx=(50,2))

r3_fs2 = Combobox(window, width=fs2_col_width)
r3_fs2['values']= (vitamin_C_list)
r3_fs2.current(0) #set the selected item
r3_fs2.grid(column=1, row=6)

r3_fs3 = Combobox(window, width=fs3_col_width)
r3_fs3['values']= (vitamin_E_list)
r3_fs3.current(0) #set the selected item
r3_fs3.grid(column=2, row=6)

r3_fs4 = Combobox(window, width=fs4_col_width)
r3_fs4['values']= (vitamin_D_list)
r3_fs4.current(0) #set the selected item
r3_fs4.grid(column=3, row=6)

r3_ws1 = Combobox(window, width=ws1_col_width)
r3_ws1['values']= (vitamin_list)
r3_ws1.current(0) #set the selected item
r3_ws1.grid(column=4, row=6)

r3_ws2 = Combobox(window, width=ws2_col_width)
r3_ws2['values']= (vitamin_list)
r3_ws2.current(0) #set the selected item
r3_ws2.grid(column=5, row=6)

r3_ws3 = Combobox(window, width=ws3_col_width)
r3_ws3['values']= (vitamin_list)
r3_ws3.current(0) #set the selected item
r3_ws3.grid(column=6, row=6)

r3_ws4 = Combobox(window, width=ws4_col_width)
r3_ws4['values']= (vitamin_list)
r3_ws4.current(0) #set the selected item
r3_ws4.grid(column=7, row=6)

r3_ws5 = Combobox(window, width=ws5_col_width)
r3_ws5['values']= (vitamin_list)
r3_ws5.current(0) #set the selected item
r3_ws5.grid(column=8, row=6)

# make fourth set of drop downs
r4_fs1 = Combobox(window, width=fs1_col_width)
r4_fs1['values']= (vitamin_A_list)
r4_fs1.current(0) #set the selected item
r4_fs1.grid(column=0, row=8, padx=(50,2))

r4_fs2 = Combobox(window, width=fs2_col_width)
r4_fs2['values']= (vitamin_C_list)
r4_fs2.current(0) #set the selected item
r4_fs2.grid(column=1, row=8)

r4_fs3 = Combobox(window, width=fs3_col_width)
r4_fs3['values']= (vitamin_E_list)
r4_fs3.current(0) #set the selected item
r4_fs3.grid(column=2, row=8)

r4_fs4 = Combobox(window, width=fs4_col_width)
r4_fs4['values']= (vitamin_D_list)
r4_fs4.current(0) #set the selected item
r4_fs4.grid(column=3, row=8)

r4_ws1 = Combobox(window, width=ws1_col_width)
r4_ws1['values']= (vitamin_list)
r4_ws1.current(0) #set the selected item
r4_ws1.grid(column=4, row=8)

r4_ws2 = Combobox(window, width=ws2_col_width)
r4_ws2['values']= (vitamin_list)
r4_ws2.current(0) #set the selected item
r4_ws2.grid(column=5, row=8)

r4_ws3 = Combobox(window, width=ws3_col_width)
r4_ws3['values']= (vitamin_list)
r4_ws3.current(0) #set the selected item
r4_ws3.grid(column=6, row=8)

r4_ws4 = Combobox(window, width=ws4_col_width)
r4_ws4['values']= (vitamin_list)
r4_ws4.current(0) #set the selected item
r4_ws4.grid(column=7, row=8)

r4_ws5 = Combobox(window, width=ws5_col_width)
r4_ws5['values']= (vitamin_list)
r4_ws5.current(0) #set the selected item
r4_ws5.grid(column=8, row=8)

# make fifth set of dropdowns
r5_fs1 = Combobox(window, width=fs1_col_width)
r5_fs1['values']= (vitamin_A_list)
r5_fs1.current(0) #set the selected item
r5_fs1.grid(column=0, row=10, padx=(50,2))

r5_fs2 = Combobox(window, width=fs2_col_width)
r5_fs2['values']= (vitamin_C_list)
r5_fs2.current(0) #set the selected item
r5_fs2.grid(column=1, row=10)

r5_fs3 = Combobox(window, width=fs3_col_width)
r5_fs3['values']= (vitamin_E_list)
r5_fs3.current(0) #set the selected item
r5_fs3.grid(column=2, row=10)

r5_fs4 = Combobox(window, width=fs4_col_width)
r5_fs4['values']= (vitamin_D_list)
r5_fs4.current(0) #set the selected item
r5_fs4.grid(column=3, row=10)

r5_ws1 = Combobox(window, width=ws1_col_width)
r5_ws1['values']= (vitamin_list)
r5_ws1.current(0) #set the selected item
r5_ws1.grid(column=4, row=10)

r5_ws2 = Combobox(window, width=ws2_col_width)
r5_ws2['values']= (vitamin_list)
r5_ws2.current(0) #set the selected item
r5_ws2.grid(column=5, row=10)

r5_ws3 = Combobox(window, width=ws3_col_width)
r5_ws3['values']= (vitamin_list)
r5_ws3.current(0) #set the selected item
r5_ws3.grid(column=6, row=10)

r5_ws4 = Combobox(window, width=ws4_col_width)
r5_ws4['values']= (vitamin_list)
r5_ws4.current(0) #set the selected item
r5_ws4.grid(column=7, row=10)

r5_ws5 = Combobox(window, width=ws5_col_width)
r5_ws5['values']= (vitamin_list)
r5_ws5.current(0) #set the selected item
r5_ws5.grid(column=8, row=10)

# sixth set of dropdowns
r6_fs1 = Combobox(window, width=fs1_col_width)
r6_fs1['values']= (vitamin_A_list)
r6_fs1.current(0) #set the selected item
r6_fs1.grid(column=0, row=12, padx=(50,2))

r6_fs2 = Combobox(window, width=fs2_col_width)
r6_fs2['values']= (vitamin_C_list)
r6_fs2.current(0) #set the selected item
r6_fs2.grid(column=1, row=12)

r6_fs3 = Combobox(window, width=fs3_col_width)
r6_fs3['values']= (vitamin_E_list)
r6_fs3.current(0) #set the selected item
r6_fs3.grid(column=2, row=12)

r6_fs4 = Combobox(window, width=fs4_col_width)
r6_fs4['values']= (vitamin_D_list)
r6_fs4.current(0) #set the selected item
r6_fs4.grid(column=3, row=12)

r6_ws1 = Combobox(window, width=ws1_col_width)
r6_ws1['values']= (vitamin_list)
r6_ws1.current(0) #set the selected item
r6_ws1.grid(column=4, row=12)

r6_ws2 = Combobox(window, width=ws2_col_width)
r6_ws2['values']= (vitamin_list)
r6_ws2.current(0) #set the selected item
r6_ws2.grid(column=5, row=12)

r6_ws3 = Combobox(window, width=ws3_col_width)
r6_ws3['values']= (vitamin_list)
r6_ws3.current(0) #set the selected item
r6_ws3.grid(column=6, row=12)

r6_ws4 = Combobox(window, width=ws4_col_width)
r6_ws4['values']= (vitamin_list)
r6_ws4.current(0) #set the selected item
r6_ws4.grid(column=7, row=12)

r6_ws5 = Combobox(window, width=ws5_col_width)
r6_ws5['values']= (vitamin_list)
r6_ws5.current(0) #set the selected item
r6_ws5.grid(column=8, row=12)

# seventh set of dropdowns
r7_fs1 = Combobox(window, width=fs1_col_width)
r7_fs1['values']= (vitamin_A_list)
r7_fs1.current(0) #set the selected item
r7_fs1.grid(column=0, row=14, padx=(50,2))

r7_fs2 = Combobox(window, width=fs2_col_width)
r7_fs2['values']= (vitamin_C_list)
r7_fs2.current(0) #set the selected item
r7_fs2.grid(column=1, row=14)

r7_fs3 = Combobox(window, width=fs3_col_width)
r7_fs3['values']= (vitamin_E_list)
r7_fs3.current(0) #set the selected item
r7_fs3.grid(column=2, row=14)

r7_fs4 = Combobox(window, width=fs4_col_width)
r7_fs4['values']= (vitamin_D_list)
r7_fs4.current(0) #set the selected item
r7_fs4.grid(column=3, row=14)

r7_ws1 = Combobox(window, width=ws1_col_width)
r7_ws1['values']= (vitamin_list)
r7_ws1.current(0) #set the selected item
r7_ws1.grid(column=4, row=14)

r7_ws2 = Combobox(window, width=ws2_col_width)
r7_ws2['values']= (vitamin_list)
r7_ws2.current(0) #set the selected item
r7_ws2.grid(column=5, row=14)

r7_ws3 = Combobox(window, width=ws3_col_width)
r7_ws3['values']= (vitamin_list)
r7_ws3.current(0) #set the selected item
r7_ws3.grid(column=6, row=14)

r7_ws4 = Combobox(window, width=ws4_col_width)
r7_ws4['values']= (vitamin_list)
r7_ws4.current(0) #set the selected item
r7_ws4.grid(column=7, row=14)

r7_ws5 = Combobox(window, width=ws5_col_width)
r7_ws5['values']= (vitamin_list)
r7_ws5.current(0) #set the selected item
r7_ws5.grid(column=8, row=14)

#--------- Bindings and Methods -----------#

get_focus_list = [
	['.!frame.!entry', '.!frame.!entry2', '.!frame.!entry3', '.!frame.!entry4', '.!frame.!combobox', '.!frame.!combobox2', '.!frame.!entry5', '.!frame.!entry6'],
	['.!frame.!entry7', '.!frame.!entry8', '.!frame.!entry9', '.!frame.!entry10', '.!frame.!combobox3', '.!frame.!combobox4', '.!frame.!entry11', '.!frame.!entry12'],
	['.!frame.!entry13', '.!frame.!entry14', '.!frame.!entry15', '.!frame.!entry16', '.!frame.!combobox5', '.!frame.!combobox6', '.!frame.!entry17', '.!frame.!entry18'],
	['.!frame.!entry19', '.!frame.!entry20', '.!frame.!entry21', '.!frame.!entry22', '.!frame.!combobox7', '.!frame.!combobox8', '.!frame.!entry23', '.!frame.!entry24'],
	['.!frame.!entry25', '.!frame.!entry26', '.!frame.!entry27', '.!frame.!entry28', '.!frame.!combobox9', '.!frame.!combobox10', '.!frame.!entry29', '.!frame.!entry30'],
	['.!frame.!entry31', '.!frame.!entry32', '.!frame.!entry33', '.!frame.!entry34', '.!frame.!combobox11', '.!frame.!combobox12', '.!frame.!entry35', '.!frame.!entry36'],
	['.!frame.!entry37', '.!frame.!entry38', '.!frame.!entry39', '.!frame.!entry40', '.!frame.!combobox13', '.!frame.!combobox14', '.!frame.!entry41', '.!frame.!entry42']
]

set_focus_list = [
	[wip1, lot1, desc1, lr1, st1, polarity1, wt1, ups1],
	[wip2, lot2, desc2, lr2, st2, polarity2, wt2, ups2],
	[wip3, lot3, desc3, lr3, st3, polarity3, wt3, ups3],
	[wip4, lot4, desc4, lr4, st4, polarity4, wt4, ups4],
	[wip5, lot5, desc5, lr5, st5, polarity5, wt5, ups5],
	[wip6, lot6, desc6, lr6, st6, polarity6, wt6, ups6],
	[wip7, lot7, desc7, lr7, st7, polarity7, wt7, ups7],
]

def map_focus(listA=get_focus_list, listB=set_focus_list):
	focus = str(root.focus_get())
	get_dict = dict((j,(x,y)) for x, i in enumerate(listA) for y, j in enumerate(i))
	set_dict = dict((j,(x,y)) for x, i in enumerate(listB) for y, j in enumerate(i))
	row = get_dict[focus][0]
	col = get_dict[focus][1]
	return row, col


# Navigation
def navigation(event):
	# This function will serve to help users navigate easily through the UPLC form maker menu.
	# A known issue is the right and left arrows will ALWAYS navigate away from a selected entry box if pressed.
	# This is a problem if the user wants to navigate within the entry box (say to edit text within that box which can only be done using the backspace key at the momment.)
	focus = str(root.focus_get())

	get_focus_list = [
		['.!frame.!entry', '.!frame.!entry2', '.!frame.!entry3', '.!frame.!entry4', '.!frame.!combobox', '.!frame.!combobox2', '.!frame.!entry5', '.!frame.!entry6'],
		['.!frame.!entry7', '.!frame.!entry8', '.!frame.!entry9', '.!frame.!entry10', '.!frame.!combobox3', '.!frame.!combobox4', '.!frame.!entry11', '.!frame.!entry12'],
		['.!frame.!entry13', '.!frame.!entry14', '.!frame.!entry15', '.!frame.!entry16', '.!frame.!combobox5', '.!frame.!combobox6', '.!frame.!entry17', '.!frame.!entry18'],
		['.!frame.!entry19', '.!frame.!entry20', '.!frame.!entry21', '.!frame.!entry22', '.!frame.!combobox7', '.!frame.!combobox8', '.!frame.!entry23', '.!frame.!entry24'],
		['.!frame.!entry25', '.!frame.!entry26', '.!frame.!entry27', '.!frame.!entry28', '.!frame.!combobox9', '.!frame.!combobox10', '.!frame.!entry29', '.!frame.!entry30'],
		['.!frame.!entry31', '.!frame.!entry32', '.!frame.!entry33', '.!frame.!entry34', '.!frame.!combobox11', '.!frame.!combobox12', '.!frame.!entry35', '.!frame.!entry36'],
		['.!frame.!entry37', '.!frame.!entry38', '.!frame.!entry39', '.!frame.!entry40', '.!frame.!combobox13', '.!frame.!combobox14', '.!frame.!entry41', '.!frame.!entry42']
	]

	# Flatten 2D list to 1D list
	all_focus_list = []
	for list in get_focus_list:
		for cell in list:
			all_focus_list.append(cell)

	# Iterate over 1D list checking to see if current focus is in the list
	# If focus not in list, do nothing.
	if focus not in all_focus_list:
		return

	get_dict = dict( (j,(x,y)) for x, i in enumerate(get_focus_list) for y, j in enumerate(i))

	set_focus_list = [
		[wip1, lot1, desc1, lr1, st1, polarity1, wt1, ups1],
		[wip2, lot2, desc2, lr2, st2, polarity2, wt2, ups2],
		[wip3, lot3, desc3, lr3, st3, polarity3, wt3, ups3],
		[wip4, lot4, desc4, lr4, st4, polarity4, wt4, ups4],
		[wip5, lot5, desc5, lr5, st5, polarity5, wt5, ups5],
		[wip6, lot6, desc6, lr6, st6, polarity6, wt6, ups6],
		[wip7, lot7, desc7, lr7, st7, polarity7, wt7, ups7],
	]

	set_dict = dict( (j,(x,y)) for x, i in enumerate(set_focus_list) for y, j in enumerate(i))

	if event.keysym == "Down":
		#print(root.focus_get())
		row = get_dict[focus][0]
		col = get_dict[focus][1]
		if row == 6:
			row = 0
		else:
			row += 1
		set_focus_list[row][col].focus_set()
	if event.keysym == "Up":
		#print(root.focus_get())
		row = get_dict[focus][0]
		col = get_dict[focus][1]
		if row == 0:
			row = 6
		else:
			row -= 1
		set_focus_list[row][col].focus_set()
	if event.keysym == "Left":
		#print(root.focus_get())
		row = get_dict[focus][0]
		col = get_dict[focus][1]
		if set_focus_list[row][col].get() != '':
			pass
		else:
			if col == 0:
				col = 7
			else:
				col -= 1
			set_focus_list[row][col].focus_set()
	if event.keysym == "Right":
		#print(root.focus_get())
		row = get_dict[focus][0]
		col = get_dict[focus][1]
		if set_focus_list[row][col].get() != '':
			pass
		else:
			if col == 7:
				col = 0
			else:
				col += 1
			set_focus_list[row][col].focus_set()

# Navigation Bindings
root.bind("<Right>", navigation)
root.bind("<Left>", navigation)
root.bind("<Up>", navigation)
root.bind("<Down>", navigation)

# Copy Down function
def copy_down(event):
	# make 1D array from 2D array
	get_focus_list_1d = []
	for i, each in enumerate(get_focus_list):
		for item in get_focus_list[i]:
			get_focus_list_1d.append(item)

	current_focus = str(root.focus_get())
	if (current_focus in get_focus_list_1d) == False:
		pass
	elif (current_focus in get_focus_list_1d) == True:
		row = map_focus()[0]
		col = map_focus()[1]
		info_to_copy = str(set_focus_list[row][col].get())
		number_of_rows = len(set_focus_list)
		countdown_row = number_of_rows-1
		while row < countdown_row:
			# if entry box
			if current_focus.split('.!frame.!')[1].split('y')[0] == 'entr':
				set_focus_list[countdown_row][col].insert(0, info_to_copy)
			#if combobox
			if current_focus.split('.!frame.!')[1].split('x')[0] == 'combobo':
				# check the values list and get index of current selection
				list_of_values = set_focus_list[row][col]['values']
				index_of_current_selection = list_of_values.index(info_to_copy)
				# set combobox with correct index
				set_focus_list[countdown_row][col].current(index_of_current_selection)
			# if checkbutton
			# if current_focus.split('.!frame.!')[1].split('n')[0] == 'checkbutto':
			# 	"If you are interested in being able to press <Ctrl> + D to auto populate the rest of the check buttons please let Richard Pearson know."
			# 	pass
			countdown_row -= 1
	else:
		pass

root.bind("<Control-KeyPress-d>", copy_down)

#----------END-OF-GUI-------------#

def read_encrypted_excel(filepath, password='PASSWORD'):
	# The specification database needs to be password protected.
	# This allows the password portection to exist and still be readable.
	temp = io.BytesIO()

	with open(filepath, 'rb') as f:
		excel = msoffcrypto.OfficeFile(f)
		excel.load_key(password)
		excel.decrypt(temp)

	df = pd.read_excel(temp, engine='openpyxl', dtype=str)
	del temp
	return df

def get_specification(wip, test_list, stability_status, silent=True, R_drive_parent_folder=R_drive_parent_folder):
	try:
		#df = pd.read_excel(os.path.join(R_drive_parent_folder, 'specification_database.xlsx'), engine='openpyxl', dtype=str)
		#df = read_encrypted_excel(os.path.normpath('R:\\QC\\Laboratory\\7. Lab improvement\\Raw Material Testing Plan\\test\\password_protected_specification_database.xlsx'))
		df = read_encrypted_excel(os.path.join(R_drive_parent_folder, 'specification_database.xlsx'))
	except:
		print(f"{Fore.YELLOW}Could not access {os.path.join(str(R_drive_parent_folder), 'specification_database.xlsx')} in the R: drive. Trying local location now.{Fore.RESET}")
		#Read the xlsx file containing the specifications for UPLC testing
		df = pd.read_excel(os.path.join(os.getcwd(), 'specification_database.xlsx'), engine='openpyxl', dtype=str)

	# Make a new data frame that only contains WIP equal to the entered WIP value in the GUI
	df2 = df.loc[df['WIP/Item'] == str(wip)]

	# Reset the index values so that we can reliably pick the 0th row (first raw values)
	df3 = df2.reset_index(inplace=False, drop=True)

	# if a matching WIP was found
	if df3.empty == False:

		spec_list = []
		form_list = []
		method_list = []

		for test in test_list:
			if test != '':
				analyte = test.split(' (')[0]
				if stability_status == False:
					spec_column = analyte + ' Specification'
				elif stability_status == True:
					spec_column = analyte + ' Stability Specification'
				else:
					print(f"{Fore.RED}The program does not seem to be reading the stability checkbox correctly.\nSee the get_specification function.{Fore.RESET}")

				method = analyte + ' SOP'
				form = analyte + ' Form'

				spec = df3.iloc[0][spec_column]
				spec_sop = df3.iloc[0][method]
				vitamer = df3.iloc[0][form]

				if type(spec) != str:
					spec = ''
					if silent == False:
						print(f"{Fore.RED}No specification found in database.{Fore.RESET}")
				spec_list.append(spec)

				if type(vitamer) != str:
					vitamer = analyte
					if silent == False:
						print(f"{Fore.YELLOW}No specific vitamin form found in database.{Fore.RESET} Defaulting to ambiguous form.")
				form_list.append(vitamer)

				if type(spec_sop) != str:
					spec_sop = ''
					if silent == False:
						print(f"{Fore.RED}The SOP ID is not found in the database for {analyte}.{Fore.RESET} Please double check the specification_database.xlsx file.")
				method_list.append(spec_sop)

			else:
				spec = ''
				spec_list.append(spec)

				vitamer = ''
				form_list.append(vitamer)

				spec_sop = ''
				method_list.append(spec_sop)

		# If piece weight or usage units are empty
		if df3.iloc[0]['Piece Weight/Usage'] == None:
			if silent== False:
				print(f"{Fore.RED}Missing data in Piece Weight/Usage column in specification database.{Fore.RESET}")
			piece_weight_or_usage = ''
		elif df3.iloc[0]['PW/Usage Units'] == None:
			if silent== False:
				print(f"{Fore.RED}Missing data in PW/Usage Units column in specification database.{Fore.RESET}")
			piece_weight_or_usage = ''
		elif df3.iloc[0]['Piece Weight/Usage'] == '':
			if silent== False:
				print(f"{Fore.RED}Missing data in Piece Weight/Usage column in specification database.{Fore.RESET}")
			piece_weight_or_usage = ''
		elif df3.iloc[0]['PW/Usage Units'] == '':
			if silent== False:
				print(f"{Fore.RED}Missing data in PW/Usage Units column in specification database.{Fore.RESET}")
			piece_weight_or_usage = ''
		else:
			piece_weight_or_usage = str(df3.iloc[0]['Piece Weight/Usage']) + ' ' + df3.iloc[0]['PW/Usage Units']

		# Build the output dictionary
		specification_info = dict()
		specification_info['spec'] = spec_list
		specification_info['form'] = form_list
		specification_info['method'] = method_list
		specification_info['Sample Description'] = df3.iloc[0]['Description']
		specification_info['Sample Type'] = df3.iloc[0]['Sample Type']
		specification_info['Piece or Usage'] = piece_weight_or_usage
		specification_info['Spec Serving'] = df3.iloc[0]['spec serving']

	# If a matching WIP was not found in the specification xlsx file
	else:
		if silent == False:
			print(f'{Fore.RED}WIP/Item {wip} does not have specification data entered into the specification database.{Fore.RESET}')

		spec_list = []
		form_list = []
		method_list = []

		for test in test_list:
			# if test != '':
			spec = ''
			spec_list.append(spec)

			vitamer = ''
			form_list.append(vitamer)

			spec_sop = ''
			method_list.append(spec_sop)


		# Build the output dictionary
		specification_info = dict()
		specification_info['spec'] = spec_list
		specification_info['form'] = form_list
		specification_info['method'] = method_list
		specification_info['Sample Description'] = ''
		specification_info['Sample Type'] = ''
		specification_info['Piece or Usage'] = ''
		specification_info['Spec Serving'] = ''

	return specification_info

def click_and_exit():

	colorama.init(autoreset=False)
	print(Fore.GREEN + "  _____   __  ____   " + Fore.CYAN + "    ____     ___  _____ ______    __   ___        ")
	print(Fore.GREEN + " / ___/  /  ]|    \  " + Fore.CYAN + "   |    \   /  _]/ ___/|      T  /  ] /   \       ")
	print(Fore.GREEN + "(   \_  /  / |  _  Y " + Fore.CYAN + "   |  o  ) /  [_(   \_ |      | /  / Y     Y      ")
	print(Fore.GREEN + " \__  T/  /  |  |  | " + Fore.CYAN + "   |     TY    _]\__  Tl_j  l_j/  /  |  O  |      ")
	print(Fore.GREEN + " /  \ /   \_ |  |  | " + Fore.CYAN + "   |  O  ||   [_ /  \ |  |  | /   \_ |     |      ")
	print(Fore.GREEN + " \    \     ||  |  | " + Fore.CYAN + "   |     ||     T\    |  |  | \     |l     !      ")
	print(Fore.GREEN + "  \___j\____jl__j__j " + Fore.CYAN + "   l_____jl_____j \___j  l__j  \____j \___/       ")
	print("                                                                                                     ")
	print(Fore.YELLOW + " _____   ___   ____   ___ ___      ___ ___   ____  __  _    ___  ____                  ")
	print(Fore.YELLOW + "|     | /   \ |    \ |   T   T    |   T   T /    T|  l/ ]  /  _]|    \                 ")
	print(Fore.YELLOW + "|   __jY     Y|  D  )| _   _ |    | _   _ |Y  o  ||  ' /  /  [_ |  D  )                ")
	print(Fore.YELLOW + "|  l_  |  O  ||    / |  \_/  |    |  \_/  ||     ||    \ Y    _]|    /                 ")
	print(Fore.YELLOW + "|   _] |     ||    \ |   |   |    |   |   ||  _  ||     Y|   [_ |    \                 ")
	print(Fore.YELLOW + "|  T   l     !|  .  Y|   |   |    |   |   ||  |  ||  .  ||     T|  .  Y                ")
	print(Fore.YELLOW + "l__j    \___/ l__j\_jl___j___j    l___j___jl__j__jl__j\_jl_____jl__j\_j                ")
	print(Fore.RESET + "                                                                                        ")

	sop_info_default_input_data = {
				"QAL1203":{"std_inj_vol":10, "samp_inj_vol":10, "volume":200},   # New cyanocobalamin
				"MTH-0035":{"std_inj_vol":10, "samp_inj_vol":10, "volume":200},  # New melatonin
				"QAL1014":{"std_inj_vol":'', "samp_inj_vol":'', "volume":100},   # Old Melatonin
				"QAL0937":{"std_inj_vol":2, "samp_inj_vol":2, "volume":100},     # Folic acid method
				"QAL0904":{"std_inj_vol":2, "samp_inj_vol":'', "volume":100},    # Vitamin A acetate
				"QAL0900":{"std_inj_vol":2, "samp_inj_vol":'', "volume":100},    # Vitamin E not correct!
				"QAL0902":{"std_inj_vol":5, "samp_inj_vol":'', "volume":100},    # Vitamin D method
				"QAL0933":{"std_inj_vol":2, "samp_inj_vol":'', "volume":100},    # Biotin
				"QAL0901":{"std_inj_vol":5, "samp_inj_vol":'', "volume":1000},   # B6 and P5P inj_vol = 3 is (fixed in code below)
				"QAL0930":{"std_inj_vol":5, "samp_inj_vol":'', "volume":1000},   # B1, B2, B5, B6, B3 has inj_vol = 1 (fixed in code below)
				"QAL1201":{"std_inj_vol":2, "samp_inj_vol":'', "volume":1000},   # Caffeine
				"QAL1020":{"std_inj_vol":'', "samp_inj_vol":'', "volume":''},    # CoQ10
				"QAL0940":{"std_inj_vol":'', "samp_inj_vol":'', "volume":''},    # Chlor Acids
				"QAL0939":{"std_inj_vol":'', "samp_inj_vol":'', "volume":''},    # Curcumin
				"QAL0910":{"std_inj_vol":'', "samp_inj_vol":'', "volume":''},    # Vitamin C Titration Method
				"QAL1209":{"std_inj_vol":'', "samp_inj_vol":'', "volume":''},    # Ascorbic acid (UPLC Method)
				"QAL1202":{"std_inj_vol":'', "samp_inj_vol":'', "volume":''},    # 5-MTF
				"QAL1016":{"std_inj_vol":'', "samp_inj_vol":'', "volume":''},    # Old cyanocobalamin
				"QAL1022":{"std_inj_vol":'', "samp_inj_vol":'', "volume":''}     # Old methylcobalamin
				}

	# Function to write a pickle file
	def write_batch_number_pickle(pickle_path):
		# Open a pickle file for writing
		# if file exists, read pickle file
		if os.path.exists(pickle_path) == True:
			with open(pickle_path, 'rb') as file:
				data = pickle.load(file)
		else:
			data = {'batch_id_int': 0, 'batch_id_string': '00000'}

		new_batch_id_int = data['batch_id_int'] + 1
		with open(pickle_path, 'wb') as pickle_file:
			data = {'batch_id_int': new_batch_id_int}
			pickle.dump(data, pickle_file)

	# Function to read the batch number pickle file
	def read_batch_number_pickle(pickle_path):
		# check if file exists
		if os.path.exists(pickle_path) == False:
			# create a pickle file with batch_id = 0
			print(f'{Fore.YELLOW}No pickle file found!{Fore.RESET} Creating one now.')
			write_batch_number_pickle(pickle_path)
			batch_id_dict = {'batch_id_int': 0, 'batch_id_string': '00000'}
			return batch_id_dict
		else:
			with open(pickle_path, 'rb') as pickle_file:
				data = pickle.load(pickle_file)

			batch_id = str(data['batch_id_int'])
			batch_id_int = int(batch_id)

			if int(batch_id) < 10000:
				# any number from 0 - 9999 will have 0s padding in front
				batch_id_string = batch_id.zfill(5)
			else:
				batch_id_string = batch_id

			batch_id_dict = {'batch_id_int': batch_id_int, 'batch_id_string': batch_id_string}

			# return dictionary with 'batch_id_int' and 'batch_id_string' keys
			return batch_id_dict

	# read the pickle file (create one if needed)
	pickle_path = os.path.join(R_drive_parent_folder, 'uplc_batch_num.pickle')
	batch_id_dict = read_batch_number_pickle(pickle_path)
	batch_id_string = batch_id_dict['batch_id_string']
	batch_id_int = batch_id_dict['batch_id_int']
	print(f'Batch ID: {batch_id_string}')

	# update (add 1 to batch id) and write the pickle file
	write_batch_number_pickle(pickle_path)

	sop_info = yaml_data_xfer(filename='sops.yaml', primary_path=os.path.join(os.getcwd(), 'config'), backup_path=os.path.join(os.getcwd(), 'config'), data=sop_info_default_input_data)
	user_preferences_config = yaml_data_xfer(filename='user_preferences_config.yaml', primary_path=backup_path, backup_path='', data=user_preferences_default_data_input, silent=True)
	notebook_info_config = yaml_data_xfer(filename='notebook_info_config.yaml', primary_path=backup_path, backup_path='', data=standard_notebook_info, silent=True)

	#Create empty lists
	wips = []
	lots = []
	sample_descriptions = []
	lab_report_ids = []
	sample_types = []
	polarities = []
	weights = []
	units_per_spec_list = []
	triplicate_status_list = []
	stability_status_list = []
	planned_testing_list = [[], [], [], [], [], [], []]
	water_soluble_testing_list = [[], [], [], [], [], [], []]

	#append first set of vitamin test dropdowns to 'planned_testing_list'
	# append row 1 testing values
	planned_testing_list[0].append(r1_fs1.get()) # Vitamin A
	planned_testing_list[0].append(r1_fs2.get()) # Vitamin C
	planned_testing_list[0].append(r1_fs3.get()) # Vitamin E
	planned_testing_list[0].append(r1_fs4.get()) # Vitamin D
	planned_testing_list[0].append(r1_ws1.get()) # Water Soluble Vitamin 1
	planned_testing_list[0].append(r1_ws2.get()) # Water Soluble Vitamin 2
	planned_testing_list[0].append(r1_ws3.get()) # Water Soluble Vitamin 3
	planned_testing_list[0].append(r1_ws4.get()) # Water Soluble Vitamin 4
	planned_testing_list[0].append(r1_ws5.get()) # Water Soluble Vitamin 5

	for x in range (4, 9):
		if planned_testing_list[0][x] != "":
			water_soluble_testing_list[0].append(planned_testing_list[0][x])

	# append row 2 testing values
	planned_testing_list[1].append(r2_fs1.get()) # Vitamin A
	planned_testing_list[1].append(r2_fs2.get()) # Vitamin C
	planned_testing_list[1].append(r2_fs3.get()) # Vitamin E
	planned_testing_list[1].append(r2_fs4.get()) # Vitamin D
	planned_testing_list[1].append(r2_ws1.get()) # Water Soluble Vitamin 1
	planned_testing_list[1].append(r2_ws2.get()) # Water Soluble Vitamin 2
	planned_testing_list[1].append(r2_ws3.get()) # Water Soluble Vitamin 3
	planned_testing_list[1].append(r2_ws4.get()) # Water Soluble Vitamin 4
	planned_testing_list[1].append(r2_ws5.get()) # Water Soluble Vitamin 5

	for x in range (4, 9):
		if planned_testing_list[1][x] != "":
			water_soluble_testing_list[1].append(planned_testing_list[1][x])

	# append row 3 testing values
	planned_testing_list[2].append(r3_fs1.get()) # Vitamin A
	planned_testing_list[2].append(r3_fs2.get()) # Vitamin C
	planned_testing_list[2].append(r3_fs3.get()) # Vitamin E
	planned_testing_list[2].append(r3_fs4.get()) # Vitamin D
	planned_testing_list[2].append(r3_ws1.get()) # Water Soluble Vitamin 1
	planned_testing_list[2].append(r3_ws2.get()) # Water Soluble Vitamin 2
	planned_testing_list[2].append(r3_ws3.get()) # Water Soluble Vitamin 3
	planned_testing_list[2].append(r3_ws4.get()) # Water Soluble Vitamin 4
	planned_testing_list[2].append(r3_ws5.get()) # Water Soluble Vitamin 5

	for x in range (4, 9):
			if planned_testing_list[2][x] != "":
				water_soluble_testing_list[2].append(planned_testing_list[2][x])

	# append row 4 testing values
	planned_testing_list[3].append(r4_fs1.get()) # Vitamin A
	planned_testing_list[3].append(r4_fs2.get()) # Vitamin C
	planned_testing_list[3].append(r4_fs3.get()) # Vitamin E
	planned_testing_list[3].append(r4_fs4.get()) # Vitamin D
	planned_testing_list[3].append(r4_ws1.get()) # Water Soluble Vitamin 1
	planned_testing_list[3].append(r4_ws2.get()) # Water Soluble Vitamin 2
	planned_testing_list[3].append(r4_ws3.get()) # Water Soluble Vitamin 3
	planned_testing_list[3].append(r4_ws4.get()) # Water Soluble Vitamin 4
	planned_testing_list[3].append(r4_ws5.get()) # Water Soluble Vitamin 5

	for x in range (4, 9):
		if planned_testing_list[3][x] != "":
			water_soluble_testing_list[3].append(planned_testing_list[3][x])

	# append row 5 testing values
	planned_testing_list[4].append(r5_fs1.get()) # Vitamin A
	planned_testing_list[4].append(r5_fs2.get()) # Vitamin C
	planned_testing_list[4].append(r5_fs3.get()) # Vitamin E
	planned_testing_list[4].append(r5_fs4.get()) # Vitamin D
	planned_testing_list[4].append(r5_ws1.get()) # Water Soluble Vitamin 1
	planned_testing_list[4].append(r5_ws2.get()) # Water Soluble Vitamin 2
	planned_testing_list[4].append(r5_ws3.get()) # Water Soluble Vitamin 3
	planned_testing_list[4].append(r5_ws4.get()) # Water Soluble Vitamin 4
	planned_testing_list[4].append(r5_ws5.get()) # Water Soluble Vitamin 5

	for x in range (4, 9):
		if planned_testing_list[4][x] != "":
			water_soluble_testing_list[4].append(planned_testing_list[4][x])

	# append row 6 testing values
	planned_testing_list[5].append(r6_fs1.get()) # Vitamin A
	planned_testing_list[5].append(r6_fs2.get()) # Vitamin C
	planned_testing_list[5].append(r6_fs3.get()) # Vitamin E
	planned_testing_list[5].append(r6_fs4.get()) # Vitamin D
	planned_testing_list[5].append(r6_ws1.get()) # Water Soluble Vitamin 1
	planned_testing_list[5].append(r6_ws2.get()) # Water Soluble Vitamin 2
	planned_testing_list[5].append(r6_ws3.get()) # Water Soluble Vitamin 3
	planned_testing_list[5].append(r6_ws4.get()) # Water Soluble Vitamin 4
	planned_testing_list[5].append(r6_ws5.get()) # Water Soluble Vitamin 5

	for x in range (4, 9):
		if planned_testing_list[5][x] != "":
			water_soluble_testing_list[5].append(planned_testing_list[5][x])

	# append row 7 testing values
	planned_testing_list[6].append(r7_fs1.get()) # Vitamin A
	planned_testing_list[6].append(r7_fs2.get()) # Vitamin C
	planned_testing_list[6].append(r7_fs3.get()) # Vitamin E
	planned_testing_list[6].append(r7_fs4.get()) # Vitamin D
	planned_testing_list[6].append(r7_ws1.get()) # Water Soluble Vitamin 1
	planned_testing_list[6].append(r7_ws2.get()) # Water Soluble Vitamin 2
	planned_testing_list[6].append(r7_ws3.get()) # Water Soluble Vitamin 3
	planned_testing_list[6].append(r7_ws4.get()) # Water Soluble Vitamin 4
	planned_testing_list[6].append(r7_ws5.get()) # Water Soluble Vitamin 5

	for x in range (4, 9):
		if planned_testing_list[6][x] != "":
			water_soluble_testing_list[6].append(planned_testing_list[6][x])

	# Create message on the GUI screen
	positive_adjective_list = ['a great', 'a super', 'a good', 'an awesome', 'a productive', 'a swell', 'an excellent', 'an outstanding', 'a splendid', 'an incredible', 'a fantastic', 'a fabulous', ' an amazing']
	if name.get() != '':
		txt_on_click = f"Thank you {name.get().split()[0]}! Your UPLC forms have been made.\nHave {positive_adjective_list[random.randrange(0, len(positive_adjective_list), 1)]} day!"
	else:
		txt_on_click = f"Thank you! Your UPLC forms have been made.\nHave {positive_adjective_list[random.randrange(0, len(positive_adjective_list), 1)]} day!"
	lbl.configure(text= txt_on_click)
	time.sleep(0.1)

	# Print the UPLC field to the coresponding batch form field
	uplc_system_asset_tags = str(uplc_system_id.get())

	# appends to wips list
	if wip1.get() != '':
		wips.append(wip1.get().upper())
	if wip2.get() != '':
		wips.append(wip2.get().upper())
	if wip3.get() != '':
		wips.append(wip3.get().upper())
	if wip4.get() != '':
		wips.append(wip4.get().upper())
	if wip5.get() != '':
		wips.append(wip5.get().upper())
	if wip6.get() != '':
		wips.append(wip6.get().upper())
	if wip7.get() != '':
		wips.append(wip7.get().upper())

	# appends to lots list
	if lot1.get() != '':
		lots.append(lot1.get().upper())
	if lot2.get() != '':
		lots.append(lot2.get().upper())
	if lot3.get() != '':
		lots.append(lot3.get().upper())
	if lot4.get() != '':
		lots.append(lot4.get().upper())
	if lot5.get() != '':
		lots.append(lot5.get().upper())
	if lot6.get() != '':
		lots.append(lot6.get().upper())
	if lot7.get() != '':
		lots.append(lot7.get().upper())

	# appends to sample_descriptions list
	# if the user enters the description lower case the return will be title case where each word is capitalized.
	if desc1.get() != "":
		if desc1.get().islower() == True:
			sample_descriptions.append(desc1.get().title())
		else:
			sample_descriptions.append(desc1.get())
	else:
		sample_descriptions.append('')

	if desc2.get() != "":
		if desc2.get().islower() == True:
			sample_descriptions.append(desc2.get().title())
		else:
			sample_descriptions.append(desc2.get())
	else:
		sample_descriptions.append('')

	if desc3.get() != "":
		if desc3.get().islower() == True:
			sample_descriptions.append(desc3.get().title())
		else:
			sample_descriptions.append(desc3.get())
	else:
		sample_descriptions.append('')

	if desc4.get() != "":
		if desc4.get().islower() == True:
			sample_descriptions.append(desc4.get().title())
		else:
			sample_descriptions.append(desc4.get())
	else:
		sample_descriptions.append('')

	if desc5.get() != "":
		if desc5.get().islower() == True:
			sample_descriptions.append(desc5.get().title())
		else:
			sample_descriptions.append(desc5.get())
	else:
		sample_descriptions.append('')

	if desc6.get() != "":
		if desc6.get().islower() == True:
			sample_descriptions.append(desc6.get().title())
		else:
			sample_descriptions.append(desc6.get())
	else:
		sample_descriptions.append('')

	if desc7.get() != "":
		if desc7.get().islower() == True:
			sample_descriptions.append(desc7.get().title())
		else:
			sample_descriptions.append(desc7.get())
	else:
		sample_descriptions.append('')

	# appends to lab_report_ids list
	if lr1.get() != "":
		lab_report_ids.append(lr1.get())
	if lr2.get() != "":
		lab_report_ids.append(lr2.get())
	if lr3.get() != "":
		lab_report_ids.append(lr3.get())
	if lr4.get() != "":
		lab_report_ids.append(lr4.get())
	if lr5.get() != "":
		lab_report_ids.append(lr5.get())
	if lr6.get() != "":
		lab_report_ids.append(lr6.get())
	if lr7.get() != "":
		lab_report_ids.append(lr7.get())

	# appends to sample_types list
	if st1.get() != "":
		sample_types.append(st1.get())
	if st2.get() != "":
		sample_types.append(st2.get())
	if st3.get() != "":
		sample_types.append(st3.get())
	if st4.get() != "":
		sample_types.append(st4.get())
	if st5.get() != "":
		sample_types.append(st5.get())
	if st6.get() != "":
		sample_types.append(st6.get())
	if st7.get() != "":
		sample_types.append(st7.get())

	# appends to polarities list
	if polarity1.get() != "":
		polarities.append(polarity1.get())
	if polarity2.get() != "":
		polarities.append(polarity2.get())
	if polarity3.get() != "":
		polarities.append(polarity3.get())
	if polarity4.get() != "":
		polarities.append(polarity4.get())
	if polarity5.get() != "":
		polarities.append(polarity5.get())
	if polarity6.get() != "":
		polarities.append(polarity6.get())
	if polarity7.get() != "":
		polarities.append(polarity7.get())

	# appends to weights list
	weights.append(wt1.get())
	weights.append(wt2.get())
	weights.append(wt3.get())
	weights.append(wt4.get())
	weights.append(wt5.get())
	weights.append(wt6.get())
	weights.append(wt7.get())

	units_per_spec_list.append(ups1.get())
	units_per_spec_list.append(ups2.get())
	units_per_spec_list.append(ups3.get())
	units_per_spec_list.append(ups4.get())
	units_per_spec_list.append(ups5.get())
	units_per_spec_list.append(ups6.get())
	units_per_spec_list.append(ups7.get())

	triplicate_status_list.append(chk_tri1.get())
	triplicate_status_list.append(chk_tri2.get())
	triplicate_status_list.append(chk_tri3.get())
	triplicate_status_list.append(chk_tri4.get())
	triplicate_status_list.append(chk_tri5.get())
	triplicate_status_list.append(chk_tri6.get())
	triplicate_status_list.append(chk_tri7.get())

	stability_checkbox_list = [stability_box1, stability_box2, stability_box3, stability_box4, stability_box5, stability_box6, stability_box7]
	for stability_checkbox in stability_checkbox_list:
		stability_status_list.append(stability_checkbox.get())

	packet = io.BytesIO()

	#date and time stuff
	todays_date = date.today()
	date_string = todays_date.strftime("%m/%d/%Y")

	#delete the following line:
	current_month = todays_date.strftime("%m")
	current_day = todays_date.strftime("%d")
	current_year = todays_date.strftime("%Y")
	pdf_date = current_month + current_day + current_year

	wip_x = 90
	lot_x = 200
	sample_description_x = 305

	# Sample Prep Check Box X-positions
	sp1_x = 114
	sp2_x = 223
	sp3_x = 311
	sp4_x = 402
	sp5_x = 522
	sp6_x = 605.5
	sp7_x = 732

	sp_y = 434

	#Line 1 Y-position
	line_y = 459

	# use regex to define analyst's initials from name variable
	if name.get() != "":
		r = re.compile(r"(?:(?<=\s)|^)(?:[a-z]|\d+)", re.I)
		initials = (''.join(r.findall(name.get())))
	else:
		initials = ""
	#makes font overlays of name, date, scale, uplc system and places on the PDF
	can = canvas.Canvas(packet, pagesize=letter)

	if date_selector.get() == date_string + " (Today)":
		selected_date_string = date_string
	elif date_selector.get() == one_day_in_future_string + " (Tomorrow)":
		selected_date_string = one_day_in_future_string
	elif date_selector.get() == two_day_in_future_string + two_day_future_weekday:
		selected_date_string = two_day_in_future_string
	elif date_selector.get() == three_day_in_future_string + three_day_future_weekday:
		selected_date_string = three_day_in_future_string
	elif date_selector.get() == four_day_in_future_string + four_day_future_weekday:
		selected_date_string = four_day_in_future_string
	elif date_selector.get() == "":
		selected_date_string = ""
	else:
		selected_date_string = date_selector.get()

	can.drawString(375, 85, name.get())
	can.drawString(120, 85, selected_date_string)
	can.drawString(115, 123, scale.get())
	can.drawString(115, 152, uplc_system_asset_tags)
	can.drawString(630, 45, "Batch ID: " + batch_id_string)
	can.saveState()

	user_preferences_config = yaml_data_xfer(filename='user_preferences_config.yaml', primary_path=backup_path, backup_path='', data=user_preferences_default_data_input, silent=True)
	gmp_crossouts_font_color = user_preferences_config['GMP_crossouts_font_color']
	if len(wips) == 1:
		can.setFillColor(gmp_crossouts_font_color)
		can.setFont('Helvetica-Bold', 14)
		can.rotate(20)
		can.drawString(420, 150, initials +" N/A " + selected_date_string)
		can.restoreState()
	elif len(wips) == 2:
		can.setFillColor(gmp_crossouts_font_color)
		can.setFont('Helvetica-Bold', 14)
		can.rotate(16.5)
		can.drawString(420, 158, initials +" N/A " + selected_date_string)
		can.restoreState()
	elif len(wips) == 3:
		can.setFillColor(gmp_crossouts_font_color)
		can.setFont('Helvetica-Bold', 14)
		can.rotate(13)
		can.drawString(420, 166, initials +" N/A " + selected_date_string)
		can.restoreState()
	elif len(wips) == 4:
		can.setFillColor(gmp_crossouts_font_color)
		can.setFont('Helvetica-Bold', 14)
		can.rotate(10.5)
		can.drawString(410, 165, initials +" N/A " + selected_date_string)
		can.restoreState()
	elif len(wips) == 5:
		can.setFillColor(gmp_crossouts_font_color)
		can.setFont('Helvetica-Bold', 14)
		can.rotate(7)
		can.drawString(410, 168, initials +" N/A " + selected_date_string)
		can.restoreState()
	elif len(wips) == 6:
		can.setFillColor(gmp_crossouts_font_color)
		can.setFont('Helvetica-Bold', 14)
		can.rotate(4.5)
		can.drawString(390, 165, initials +" N/A " + selected_date_string)
		can.restoreState()

	i = 0

	while i < len(wips):

		# Dynamically set font sizes for UPLC batch sheet
		wip_font_size = set_font_size(wips[i], 'uplc_batch_wip_field') #
		lot_font_size = set_font_size(lots[i], 'uplc_batch_lot_field') #
		description_font_size = set_font_size(sample_descriptions[i], 'uplc_batch_description_field')

		# Makes UPLC Batch Sheet

		can.saveState()
		can.setFont('Helvetica', wip_font_size)
		can.drawString(wip_x, line_y, wips[i])
		can.restoreState()

		can.saveState()
		can.setFont('Helvetica', lot_font_size)
		can.drawString(lot_x, line_y, lots[i])
		can.restoreState()

		can.saveState()
		can.setFont('Helvetica', description_font_size)
		if lab_report_ids[i] != "":
			if sample_descriptions[i] != "":
				can.drawString(sample_description_x, line_y, sample_descriptions[i] + " (Lab Report ID: "+ lab_report_ids[i] + ")")
			else:
				can.drawString(sample_description_x, line_y, get_specification(wips[i], planned_testing_list[i], stability_status_list[i])['Sample Description'] + " (Lab Report ID: "+ lab_report_ids[i] + ")")
		else:
			can.drawString(sample_description_x, line_y, get_specification(wips[i], planned_testing_list[i], stability_status_list[i])['Sample Description'])
		can.restoreState()

		can.drawString(sp1_x, sp_y, "X")
		can.drawString(sp2_x, sp_y, "X")
		can.drawString(sp3_x, sp_y, "X")
		can.drawString(sp4_x, sp_y, "X")
		can.drawString(sp5_x, sp_y, "X")
		can.drawString(sp6_x, sp_y, "X")
		can.drawString(sp7_x, sp_y, "X")

		i += 1
		line_y = line_y - 43.5
		sp_y = sp_y - 43.25

	# Draw lines on the UPLC Batch Preparation Form
	batch_line_x1 = 35
	batch_line_y1 = 173
	batch_line_x2 = 757
	# the y2 will change
	if len(wips) == 1:
		can.line(batch_line_x1, batch_line_y1, batch_line_x2, 432)
	elif len(wips) == 2:
		can.line(batch_line_x1, batch_line_y1, batch_line_x2, 388)
	elif len(wips) == 3:
		can.line(batch_line_x1, batch_line_y1, batch_line_x2, 345)
	elif len(wips) == 4:
		can.line(batch_line_x1, batch_line_y1, batch_line_x2, 302)
	elif len(wips) == 5:
		can.line(batch_line_x1, batch_line_y1, batch_line_x2, 259)
	elif len(wips) == 6:
		can.line(batch_line_x1, batch_line_y1, batch_line_x2, 215)
	elif len(wips) == 7:
		pass
	else:
		pass

	can.save()
	# move to the beginning of the StringIO buffer
	packet.seek(0)
	new_pdf = PdfFileReader(packet)
	# read the existing PDF
	current_working_directory = str(os.getcwd())

	# Opens the UPLC Batch Form
	existing_pdf = PdfFileReader(open(current_working_directory + r"\forms\uplc_batch_form.pdf", "rb"))

	output = PdfFileWriter()
	# add the "watermark" (which is the new pdf) on the existing page
	page = existing_pdf.getPage(0)
	page2 = new_pdf.getPage(0)
	page.mergePage(page2)
	output.addPage(page)
	# Write "output" to a real file

	outputStream = open("new_uplc_batch_form.pdf", "wb")
	output.write(outputStream)
	outputStream.close()

	# Appends the created pdf into one pdf document for easy printing
	batch_file_name = str(name.get()) + "_" + str(pdf_date) + "_batch" + ".pdf"
	batch_file_name_no_pdf = str(name.get()) + "_" + str(pdf_date) + "_batch"

	merger = PdfFileMerger()

	merger.append(PdfFileReader("new_uplc_batch_form.pdf", 'rb'))
	os.remove("new_uplc_batch_form.pdf")

	# Makes Each UPLC Sample Sheet
	i = 0
	while i < len(wips):
		packet = io.BytesIO()
		can = canvas.Canvas(packet, pagesize=letter)

		print('\n*************************************************************************\n')
		print(f'WIP/Item: {wips[i]}')
		print(f'Sample Description: {sample_descriptions[i]}')
		print(f'Lot: {lots[i]}')
		print(f'Lab Report: {lab_report_ids[i]}')
		if triplicate_status_list[i] == True:
			print(f'{Fore.YELLOW}Triplicate Required: {Fore.GREEN}{triplicate_status_list[i]}{Fore.RESET}')
		else:
			print(f'Triplicate Required: {triplicate_status_list[i]}')
		print(f'Stability Sample: {stability_status_list[i]}\n')

		# Dictionaries containing all x, y cordinates for each row
		# water soluble form
		row_1_water_soluble_form_positions = {'volume': [313, 640], 'sop_positions': [53, 562], 'analyte': [139, 631], 'std_inj_vol':[400, 626], 'sample_inj_vol':[535, 626], 'correction_factor':[512, 605], 'acceptable_methods':[153, 562], 'notebook_number':[115, 608], 'stock_page': [130, 584], 'working_page':[260, 584]}
		row_2_water_soluble_form_positions = {'volume': [313, 542], 'sop_positions': [53, 469], 'analyte': [139, 534], 'std_inj_vol':[400, 530], 'sample_inj_vol':[535, 530], 'correction_factor':[512, 509], 'acceptable_methods':[153, 469], 'notebook_number':[115, 510], 'stock_page': [130, 487], 'working_page':[260, 487]}
		row_3_water_soluble_form_positions = {'volume': [313, 450], 'sop_positions': [53, 371], 'analyte': [139, 444], 'std_inj_vol':[400, 435], 'sample_inj_vol':[535, 435], 'correction_factor':[512, 415], 'acceptable_methods':[153, 371], 'notebook_number':[115, 419], 'stock_page': [130, 395], 'working_page':[260, 395]}
		row_4_water_soluble_form_positions = {'volume': [313, 347], 'sop_positions': [53, 268], 'analyte': [139, 344], 'std_inj_vol':[400, 335], 'sample_inj_vol':[535, 335], 'correction_factor':[512, 315], 'acceptable_methods':[153, 268], 'notebook_number':[115, 320], 'stock_page': [130, 296], 'working_page':[260, 296]}
		row_5_water_soluble_form_positions = {'volume': [313, 245], 'sop_positions': [53, 170], 'analyte': [139, 241], 'std_inj_vol':[400, 235], 'sample_inj_vol':[535, 235], 'correction_factor':[512, 215], 'acceptable_methods':[153, 170], 'notebook_number':[115, 216], 'stock_page': [130, 193], 'working_page':[260, 193]}

		# fat soluble form
		# Maps to the Vitamin A box
		row_1_fat_soluble_form_positions = {'specification_positions':[163, 626], 'volume': [315, 633], 'sop_positions': [53, 513], 'analyte': [52, 636], 'std_inj_vol':[405, 622], 'sample_inj_vol':[535, 622], 'correction_factor':[512, 602], 'acceptable_methods':[153, 513], 'notebook_number':[114, 566], 'stock_page': [127, 541], 'working_page':[260, 541]}
		# Maps to the Vitamin C box
		row_2_fat_soluble_form_positions = {'specification_positions':[163, 218], 'volume': [315, 550], 'sop_positions': [108, 195], 'analyte': [52, 237], 'std_inj_vol':[405, 600], 'sample_inj_vol':[535, 600], 'correction_factor':[512, 555], 'acceptable_methods':[158, 195], 'notebook_number':[114, 220], 'stock_page': [127, 190], 'working_page':[260, 190]}
		# Maps to the Vitamin E box
		row_3_fat_soluble_form_positions = {'specification_positions':[163, 480], 'volume': [315, 490], 'sop_positions': [53, 380], 'analyte': [52, 491], 'std_inj_vol':[405, 480], 'sample_inj_vol':[535, 480], 'correction_factor':[512, 460], 'acceptable_methods':[153, 380], 'notebook_number':[114, 429], 'stock_page': [127, 418], 'working_page':[260, 418]}
		# Maps to the Vitamin D box
		row_4_fat_soluble_form_positions = {'specification_positions':[163, 347], 'volume': [315, 357], 'sop_positions': [53, 255], 'analyte': [52, 358], 'std_inj_vol':[405, 347], 'sample_inj_vol':[535, 347], 'correction_factor':[512, 327], 'acceptable_methods':[153, 255], 'notebook_number':[114, 305], 'stock_page': [127, 294], 'working_page':[260, 294]}

		ws = {
			0:{
				r1_ws1: row_1_water_soluble_form_positions,
				r1_ws2: row_2_water_soluble_form_positions,
				r1_ws3: row_3_water_soluble_form_positions,
				r1_ws4: row_4_water_soluble_form_positions,
				r1_ws5: row_5_water_soluble_form_positions
				},
			1:{
				r2_ws1: row_1_water_soluble_form_positions,
				r2_ws2: row_2_water_soluble_form_positions,
				r2_ws3: row_3_water_soluble_form_positions,
				r2_ws4: row_4_water_soluble_form_positions,
				r2_ws5: row_5_water_soluble_form_positions
			},
			2:{
				r3_ws1: row_1_water_soluble_form_positions,
				r3_ws2: row_2_water_soluble_form_positions,
				r3_ws3: row_3_water_soluble_form_positions,
				r3_ws4: row_4_water_soluble_form_positions,
				r3_ws5: row_5_water_soluble_form_positions
			},
			3:{
				r4_ws1: row_1_water_soluble_form_positions,
				r4_ws2: row_2_water_soluble_form_positions,
				r4_ws3: row_3_water_soluble_form_positions,
				r4_ws4: row_4_water_soluble_form_positions,
				r4_ws5: row_5_water_soluble_form_positions
			},
			4:{
				r5_ws1: row_1_water_soluble_form_positions,
				r5_ws2: row_2_water_soluble_form_positions,
				r5_ws3: row_3_water_soluble_form_positions,
				r5_ws4: row_4_water_soluble_form_positions,
				r5_ws5: row_5_water_soluble_form_positions
			},
			5:{
				r6_ws1: row_1_water_soluble_form_positions,
				r6_ws2: row_2_water_soluble_form_positions,
				r6_ws3: row_3_water_soluble_form_positions,
				r6_ws4: row_4_water_soluble_form_positions,
				r6_ws5: row_5_water_soluble_form_positions
			},
			6:{
				r7_ws1: row_1_water_soluble_form_positions,
				r7_ws2: row_2_water_soluble_form_positions,
				r7_ws3: row_3_water_soluble_form_positions,
				r7_ws4: row_4_water_soluble_form_positions,
				r7_ws5: row_5_water_soluble_form_positions
			}
			}

		fs = {
			0:{
				r1_fs1: row_1_fat_soluble_form_positions,
				r1_fs2: row_2_fat_soluble_form_positions,
				r1_fs3: row_3_fat_soluble_form_positions,
				r1_fs4: row_4_fat_soluble_form_positions
				},
			1:{
				r2_fs1: row_1_fat_soluble_form_positions,
				r2_fs2: row_2_fat_soluble_form_positions,
				r2_fs3: row_3_fat_soluble_form_positions,
				r2_fs4: row_4_fat_soluble_form_positions
			},
			2:{
				r3_fs1: row_1_fat_soluble_form_positions,
				r3_fs2: row_2_fat_soluble_form_positions,
				r3_fs3: row_3_fat_soluble_form_positions,
				r3_fs4: row_4_fat_soluble_form_positions
			},
			3:{
				r4_fs1: row_1_fat_soluble_form_positions,
				r4_fs2: row_2_fat_soluble_form_positions,
				r4_fs3: row_3_fat_soluble_form_positions,
				r4_fs4: row_4_fat_soluble_form_positions
			},
			4:{
				r5_fs1: row_1_fat_soluble_form_positions,
				r5_fs2: row_2_fat_soluble_form_positions,
				r5_fs3: row_3_fat_soluble_form_positions,
				r5_fs4: row_4_fat_soluble_form_positions
			},
			5:{
				r6_fs1: row_1_fat_soluble_form_positions,
				r6_fs2: row_2_fat_soluble_form_positions,
				r6_fs3: row_3_fat_soluble_form_positions,
				r6_fs4: row_4_fat_soluble_form_positions
			},
			6:{
				r7_fs1: row_1_fat_soluble_form_positions,
				r7_fs2: row_2_fat_soluble_form_positions,
				r7_fs3: row_3_fat_soluble_form_positions,
				r7_fs4: row_4_fat_soluble_form_positions
			}
			}

		can.drawString(227, 704, name.get())
		can.drawString(88, 704, selected_date_string)
		can.drawString(125, 688, wips[i])
		can.drawString(318, 688, lots[i])
		can.drawString(125, 672, sample_descriptions[i])
		can.drawString(470, 18, 'Batch ID: ' + batch_id_string)
		page_number = str(i + 1)
		can.drawString(495, 8, 'Page ' + page_number + ' of ' + str(len(wips)))

		# Get the number of water soluble tests
		ws_test_num = 0
		for each in water_soluble_testing_list[i]:
			if each != '':
				ws_test_num += 1

		if polarities[i] == "Water Soluble":
			for idx, row in enumerate(ws[i]):
				# print(f'idx is {idx}')
				# print(f'row is {row}')
				if str(row.get()) != '': # if dropdown is not empty then write the QALnumber in the right place and the correct volume etc.
					sop = str(row.get()).split(' (')[1].split(')')[0]
					analyte = str(row.get()).split(' (')[0]

					print('------------------------------------------------------------------------')
					print(f'Planned SOP: {sop}')
					print(f'Analyte: {analyte}')
					for each in list(sop_info):
						if each == sop:
							if sop_info[each]["volume"] == '':
								volume = ''
							else:
								volume = "Volume = " + str(sop_info[each]["volume"]) + " mL"
							standard_injection_volume = str(sop_info[each]["std_inj_vol"])
							sample_injection_volume = str(sop_info[each]["samp_inj_vol"])

							try:
								injection_volume_correction_factor = str(float(sample_injection_volume) / float(standard_injection_volume))
							except:
								injection_volume_correction_factor = ""

							spec = get_specification(wips[i], planned_testing_list[i], stability_status_list[i], silent=True)['spec'][idx + 4]
							form = get_specification(wips[i], planned_testing_list[i], stability_status_list[i], silent=True)['form'][idx + 4]
							method = get_specification(wips[i], planned_testing_list[i], stability_status_list[i], silent=True)['method'][idx + 4]

							# Solution for different injection volume for B3 in QAL0930
							# Solution for different injection volume for P5P in QAL0901
							if (sop == 'QAL0930') & (analyte == 'B3'):
								standard_injection_volume = str(1)
							if (sop == 'QAL0901') & (analyte == 'P5P'):
								standard_injection_volume = str(3)

							try:
								notebook_number = str(notebook_info_config[sop]['notebook_number'])
							except:
								notebook_number = ''
							try:
								stock_page = str(notebook_info_config[sop]['stock_page'])
							except:
								stock_page = ''
							try:
								working_page = str(notebook_info_config[sop]['working_page'])
							except:
								working_page = ''

							if user_preferences_config['print_standard_notebook_info'] == True:
								can.drawString(ws[i][row]['notebook_number'][0], ws[i][row]['notebook_number'][1], notebook_number)
								can.drawString(ws[i][row]['stock_page'][0], ws[i][row]['stock_page'][1], stock_page)
								can.drawString(ws[i][row]['working_page'][0], ws[i][row]['working_page'][1], working_page)

							can.saveState()
							can.setFont('Helvetica', 8)
							can.drawString(ws[i][row]['volume'][0], ws[i][row]['volume'][1], volume)
							can.drawString(ws[i][row]['sop_positions'][0], ws[i][row]['sop_positions'][1], sop)
							can.restoreState()

							can.saveState()
							can.setFont('Helvetica-Oblique', 8)
							can.drawString(ws[i][row]['acceptable_methods'][0], ws[i][row]['acceptable_methods'][1], method)
							can.restoreState()

							if (type(spec) != str) or (spec == ''):
								spec = ''
								print(f'Form: {form}')
								print(f'Spec: {spec}')
								print(f'Acceptable Methods: {method}')
								print(f'\n{Fore.RED}WARNING:{Fore.RESET}\n{analyte} specification for WIP/Item {wips[i]} not found in specification database!\n')

							if (type(form) != str) or (form == ''):
								form = analyte
								print(f"{Fore.YELLOW}No analyte form found.{Fore.RESET} Defaulting to ambiguous vitamin form.")

							# Logic to put analyte or form on the PDF document
							if (analyte == 'B1') or (analyte == 'B2') or (analyte == 'B5'):
								spec_line = analyte + ': ' + spec
							elif analyte == 'B6':
								if 'hcl' in form.lower():
									form_modified = 'Pyridoxine'
									spec_line = form_modified + ': ' + spec
								elif form.lower() == 'pyridoxal 5-phosphate':
									print(f'form: {form}')
									form_modified = 'P5P'
									spec_line = form_modified + ': ' + spec
								else:
									spec_line = form + ': ' + spec
							else:
								spec_line = form + ': ' + spec

							if spec != '':
								print(f'Form: {form}')
								print(f'Spec: {spec}')
								print(f'Acceptable Methods: {method}')

							can.saveState()
							can.setFont('Helvetica', 8)
							can.drawString(ws[i][row]['analyte'][0], ws[i][row]['analyte'][1], spec_line)
							can.restoreState()

							if user_preferences_config['print_injection_volumes'] == True:
								can.drawString(ws[i][row]['std_inj_vol'][0], ws[i][row]['std_inj_vol'][1], standard_injection_volume)
								can.drawString(ws[i][row]['sample_inj_vol'][0], ws[i][row]['sample_inj_vol'][1], sample_injection_volume)
								can.drawString(ws[i][row]['correction_factor'][0], ws[i][row]['correction_factor'][1], injection_volume_correction_factor)

		if polarities[i] == "Fat Soluble":
			for idx, row in enumerate(fs[i]):
				# print(f'idx is {idx}')
				# print(f'row is {row}')
				if str(row.get()) != '': # if dropdown is not empty then write the QALnumber in the right place and the correct volume etc.
					sop = str(row.get()).split(' (')[1].split(')')[0]
					analyte = str(row.get()).split(' (')[0]

					print('------------------------------------------------------------------------')
					print(f'Planned SOP: {sop}')
					print(f'Analyte: {analyte}')

					for each in list(sop_info):
						if each == sop:
							if sop_info[each]["volume"] == '':
								volume = ''
							else:
								volume = "Volume = " + str(sop_info[each]["volume"]) + " mL"
							standard_injection_volume = str(sop_info[each]["std_inj_vol"])
							sample_injection_volume = str(sop_info[each]["samp_inj_vol"])
							if isinstance(sop_info[each]["std_inj_vol"], str) == False and isinstance(sop_info[each]["samp_inj_vol"], str) == False:
								injection_volume_correction_factor = str(float(sample_injection_volume) / float(standard_injection_volume))
							else:
								injection_volume_correction_factor = ''

							spec = get_specification(wips[i], planned_testing_list[i], stability_status_list[i], silent=True)['spec'][idx]
							form = get_specification(wips[i], planned_testing_list[i], stability_status_list[i], silent=True)['form'][idx]
							method = get_specification(wips[i], planned_testing_list[i], stability_status_list[i], silent=True)['method'][idx]

							try:
								notebook_number = str(notebook_info_config[sop]['notebook_number'])
							except:
								notebook_number = ''
							try:
								stock_page = str(notebook_info_config[sop]['stock_page'])
							except:
								stock_page = ''
							try:
								working_page = str(notebook_info_config[sop]['working_page'])
							except:
								working_page = ''

							if user_preferences_config['print_standard_notebook_info'] == True:
								can.drawString(fs[i][row]['notebook_number'][0], fs[i][row]['notebook_number'][1], notebook_number)
								can.drawString(fs[i][row]['stock_page'][0], fs[i][row]['stock_page'][1], stock_page)
								can.drawString(fs[i][row]['working_page'][0], fs[i][row]['working_page'][1], working_page)


							can.saveState()
							can.setFont('Helvetica', 8)
							can.drawString(fs[i][row]['volume'][0], fs[i][row]['volume'][1], volume)
							can.drawString(fs[i][row]['sop_positions'][0], fs[i][row]['sop_positions'][1], sop)
							can.restoreState()

							can.saveState()
							can.setFont('Helvetica-Oblique', 8)
							can.drawString(fs[i][row]['acceptable_methods'][0], fs[i][row]['acceptable_methods'][1], method)
							can.restoreState()

							if (type(spec) != str) or (spec == ''):
								spec = ''
								print(f'Form: {form}')
								print(f'Spec: {spec}')
								print(f'Acceptable Methods: {method}')
								print(f'\n{Fore.RED}WARNING:{Fore.RESET}\n{analyte} specification for WIP/Item {wips[i]} not found in specification database!\n')

							if (type(form) != str) or (form == ''):
								form = analyte
								print(f"{Fore.YELLOW}No analyte form found.{Fore.RESET} Defaulting to ambiguous vitamin form.")

							spec_line = form + ': ' + spec

							if spec != '':
								print(f'Form: {form}')
								print(f'Spec: {spec}')
								print(f'Acceptable Methods: {method}')

							can.saveState()
							can.setFont('Helvetica', 8)
							#can.drawString(fs[i][row]['analyte'][0], fs[i][row]['analyte'][1], spec_line)
							can.drawString(fs[i][row]['analyte'][0], fs[i][row]['analyte'][1], form)
							can.drawString(fs[i][row]['specification_positions'][0], fs[i][row]['specification_positions'][1], spec)
							can.restoreState()

							if user_preferences_config['print_injection_volumes'] == True:
								can.drawString(fs[i][row]['std_inj_vol'][0], fs[i][row]['std_inj_vol'][1], standard_injection_volume)
								can.drawString(fs[i][row]['sample_inj_vol'][0], fs[i][row]['sample_inj_vol'][1], sample_injection_volume)
								can.drawString(fs[i][row]['correction_factor'][0], fs[i][row]['correction_factor'][1], injection_volume_correction_factor)


		print('------------------------------------------------------------------------')


		if lab_report_ids[i] != '':
			can.drawString(395, 655, "Lab Report ID: " + lab_report_ids[i])

		if triplicate_status_list[i] == True:
			can.saveState()
			can.setFillColor('Red')
			can.setFont('Helvetica-Bold', 14)
			can.drawString(395, 665, "TRIPLICATE REQUIRED")
			can.restoreState()

		# if it is a water soluble finished product
		if sample_types[i] == "Finished Product" and polarities[i] != "Fat Soluble":
			can.drawString(108, 655, "X")
			can.drawString(310, 657, "N/A")
			if weights[i] != '':
				if (ws_test_num >= 1) or (triplicate_status_list[i] == True):
					can.drawString(315, 133, weights[i] + " g")
				if (ws_test_num >= 2) or (triplicate_status_list[i] == True):
					can.drawString(390, 133, weights[i] + " g")
				if (ws_test_num >= 3) or (triplicate_status_list[i] == True):
					can.drawString(470, 133, weights[i] + " g")
			if units_per_spec_list[i] != '':
				if (ws_test_num >= 1) or (triplicate_status_list[i] == True):
					can.drawString(315, 113, str(units_per_spec_list[i]))
				# next 3 lines added to fix ValueError: could not convert string to float.
				if weights[i] == '':
					spec_weight = ''
				else:
					spec_weight = round(float(weights[i]) * float(units_per_spec_list[i]), 8)
					spec_weight = str(spec_weight)
					if (ws_test_num >= 1) or (triplicate_status_list[i] == True):
						can.drawString(315, 93, spec_weight + " g")

				if (ws_test_num >= 2) or (triplicate_status_list[i] == True):
					can.drawString(390, 113, str(units_per_spec_list[i]))
				if (ws_test_num >= 3) or (triplicate_status_list[i] == True):
					can.drawString(470, 113, str(units_per_spec_list[i]))

				if (weights[i] != '') and (units_per_spec_list[i] != ''):
					spec_weight = round(float(weights[i]) * float(units_per_spec_list[i]), 8)
					spec_weight = str(spec_weight)
					if (ws_test_num >= 2) or (triplicate_status_list[i] == True):
						can.drawString(390, 93, spec_weight + " g")
					if (ws_test_num >= 3) or (triplicate_status_list[i] == True):
						can.drawString(470, 93, spec_weight + " g")

		elif sample_types[i] == "Finished Product" and polarities[i] == "Fat Soluble":
			can.drawString(108, 655, "X")
			can.drawString(310, 657, "N/A")
			if weights[i] != '':
				can.drawString(245, 153, weights[i] + " g")
			if units_per_spec_list[i] != '':
				can.drawString(245, 130, str(units_per_spec_list[i]))
				# next 3 lines added to fix ValueError: could not convert string to float.
				if weights[i] == '':
					spec_weight = ''
				else:
					spec_weight = round(float(weights[i]) * float(units_per_spec_list[i]), 8)
					spec_weight = str(spec_weight)
					can.drawString(245, 111, spec_weight + " g")

		# if it is a water soluble raw blend
		elif sample_types[i] == "Raw Blend" and polarities[i] != "Fat Soluble":
			can.drawString(198, 655, "X")
			if (ws_test_num >= 1) or (triplicate_status_list[i] == True):
				can.drawString(315, 133, "N/A")
				can.drawString(315, 110, "N/A")
			if (ws_test_num >= 2) or (triplicate_status_list[i] == True):
				can.drawString(390, 133, "N/A")
				can.drawString(390, 110, "N/A")
			if (ws_test_num >= 3) or (triplicate_status_list[i] == True):
				can.drawString(470, 110, "N/A")
				can.drawString(470, 133, "N/A")
			if weights[i] != '':
				can.drawString(310, 658, weights[i] + " mg")
				usage_rate_in_grams = round(float(weights[i]) / 1000, 8)
				usage_rate_in_grams = str(usage_rate_in_grams)
				if (ws_test_num >= 1) or (triplicate_status_list[i] == True):
					can.drawString(313, 90, usage_rate_in_grams + " g")
				if (ws_test_num >= 2) or (triplicate_status_list[i] == True):
					can.drawString(390, 90, usage_rate_in_grams + " g")
				if (ws_test_num >= 3) or (triplicate_status_list[i] == True):
					can.drawString(470, 90, usage_rate_in_grams + " g")

		elif sample_types[i] == "Raw Blend" and polarities[i] == "Fat Soluble":
			can.drawString(198, 655, "X")
			can.drawString(245, 153, "N/A")
			can.drawString(245, 133, "N/A")
			if weights[i] != '':
				can.drawString(310, 658, weights[i] + " mg")
				usage_rate_in_grams = round(float(weights[i]) / 1000, 8)
				usage_rate_in_grams = str(usage_rate_in_grams)
				can.drawString(245, 110, usage_rate_in_grams + " g")

		# if it is a water soluble 'percent active' aka pure material
		elif sample_types[i] == "Percent Active" and polarities[i] != "Fat Soluble":
			can.drawString(198, 655, "X")
			if (ws_test_num >= 1) or (triplicate_status_list[i] == True):
				can.drawString(315, 133, "N/A")
				can.drawString(315, 110, "N/A")
			if (ws_test_num >= 2) or (triplicate_status_list[i] == True):
				can.drawString(390, 133, "N/A")
				can.drawString(390, 110, "N/A")
			if (ws_test_num >= 3) or (triplicate_status_list[i] == True):
				can.drawString(470, 133, "N/A")
				can.drawString(470, 110, "N/A")
				if weights[i] == '':
					can.drawString(390, 90, "N/A")
					can.drawString(470, 90, "N/A")
			if weights[i] != '':
				can.drawString(310, 658, weights[i] + " mg")
				usage_rate_in_grams = round(float(weights[i]) / 1000, 8)
				usage_rate_in_grams = str(usage_rate_in_grams)
				if (ws_test_num >= 1) or (triplicate_status_list[i] == True):
					can.drawString(313, 90, usage_rate_in_grams + " g")
				if (ws_test_num >= 2) or (triplicate_status_list[i] == True):
					can.drawString(390, 90, usage_rate_in_grams + " g")
				if (ws_test_num >= 3) or (triplicate_status_list[i] == True):
					can.drawString(470, 90, usage_rate_in_grams + " g")
			else:
				# print 'N/A' in the top of the form
				can.drawString(310, 658, "N/A")
				# print 'N/A' in the usage rate bottom area?
				#can.drawString(313, 90, "N/A 2")

		elif sample_types[i] == "Percent Active" and polarities[i] == "Fat Soluble":
			can.drawString(198, 655, "X")
			if weights[i] != '':
				can.drawString(310, 658, weights[i] + " mg")
			else:
				can.drawString(310, 658, "N/A")
			can.drawString(245, 153, "N/A")
			can.drawString(245, 133, "N/A")
			if weights[i] != '':
				usage_rate_in_grams = round(float(weights[i]) / 1000, 8)
				usage_rate_in_grams = str(usage_rate_in_grams)
				can.drawString(245, 110, usage_rate_in_grams + " g")
			else:
				can.drawString(245, 110, "N/A")

		can.save()
		packet.seek(0)
		new_pdf = PdfFileReader(packet)

		# If statement to choose the correct cGMP form and N/A, initial, and date the forms!
		if planned_testing_list[i][0] != "" and planned_testing_list[i][1] == "" and planned_testing_list[i][2] == "" and planned_testing_list[i][3] == "": # Vitamin A only
			existing_pdf = PdfFileReader(open(r"forms\fat_soluble\uplc_sample_form_fs_a.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)
			can2.saveState()
			can2.rotate(14)
			#This is the rotated string values for the upper portion
			#can2.drawString(370, 487, initials +" N/A " + selected_date_string) # initial and date A
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			can2.drawString(325, 355, initials +" N/A " + selected_date_string) # initial and date E
			can2.drawString(310, 230, initials +" N/A " + selected_date_string) # initial and date D
			can2.restoreState()

			can2.saveState()
			can2.rotate(7)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			can2.drawString(255, 185, initials +" N/A " + selected_date_string) # initial and date C - upper box
			can2.restoreState()

			can2.saveState()
			can2.rotate(30)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			can2.drawString(380, -120, initials +" N/A " + selected_date_string) # initial and date C - lower right box
			can2.restoreState()
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			outputStream2 = open(r"forms\fat_soluble\uplc_sample_form_fs_a2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			os.remove(r"forms\fat_soluble\uplc_sample_form_fs_a2.pdf")

		elif planned_testing_list[i][0] == "" and planned_testing_list[i][1] != "" and planned_testing_list[i][2] == "" and planned_testing_list[i][3] == "": # Vitamin C only
			existing_pdf = PdfFileReader(open(r"forms\fat_soluble\uplc_sample_form_fs_c.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)
			can2.saveState()
			can2.rotate(14)
			#This is the rotated string values for the upper portion
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			can2.drawString(370, 487, initials +" N/A " + selected_date_string) # initial and date A
			can2.drawString(325, 355, initials +" N/A " + selected_date_string) # initial and date E
			can2.drawString(310, 230, initials +" N/A " + selected_date_string) # initial and date D
			can2.restoreState()

			can2.saveState()
			can2.rotate(30)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			#can2.drawString(380, -120, initials +" N/A " + selected_date_string) # initial and date C - lower right box
			can2.drawString(150, 10, initials +" N/A " + selected_date_string) # fat soluble - lower left box
			can2.restoreState()
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			outputStream2 = open(r"forms\fat_soluble\uplc_sample_form_fs_c2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			os.remove(r"forms\fat_soluble\uplc_sample_form_fs_c2.pdf")

		elif planned_testing_list[i][0] == "" and planned_testing_list[i][1] == "" and planned_testing_list[i][2] != "" and planned_testing_list[i][3] == "": # Vitamin E only
			existing_pdf = PdfFileReader(open(r"forms\fat_soluble\uplc_sample_form_fs_e.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)
			can2.saveState()
			can2.rotate(14)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			#This is the rotated string values for the upper portion
			can2.drawString(370, 487, initials +" N/A " + selected_date_string) # initial and date A
			#can2.drawString(325, 355, initials +" N/A " + selected_date_string) # initial and date E
			can2.drawString(310, 230, initials +" N/A " + selected_date_string) # initial and date D
			can2.restoreState()

			can2.saveState()
			can2.rotate(7)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			can2.drawString(255, 185, initials +" N/A " + selected_date_string) # initial and date C - upper box
			can2.restoreState()

			can2.saveState()
			can2.rotate(30)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			can2.drawString(380, -120, initials +" N/A " + selected_date_string) # initial and date C - lower right box
			#can2.drawString(150, 10, initials +" N/A " + selected_date_string) # fat soluble - lower left box
			can2.restoreState()
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			outputStream2 = open(r"forms\fat_soluble\uplc_sample_form_fs_e2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			os.remove(r"forms\fat_soluble\uplc_sample_form_fs_e2.pdf")

		# Condition if Vitamin D ONLY
		elif planned_testing_list[i][0] == "" and planned_testing_list[i][1] == "" and planned_testing_list[i][2] == "" and planned_testing_list[i][3] != "": # Vitamin D only
			existing_pdf = PdfFileReader(open(r"forms\fat_soluble\uplc_sample_form_fs_d.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)
			can2.saveState()
			can2.rotate(14)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			#This is the rotated string values for the upper portion
			can2.drawString(370, 487, initials +" N/A " + selected_date_string) # initial and date A
			can2.drawString(325, 355, initials +" N/A " + selected_date_string) # initial and date E
			#can2.drawString(310, 230, initials +" N/A " + selected_date_string) # initial and date D
			can2.restoreState()

			can2.saveState()
			can2.rotate(7)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			can2.drawString(255, 185, initials +" N/A " + selected_date_string) # initial and date C - upper box
			can2.restoreState()

			can2.saveState()
			can2.rotate(30)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			can2.drawString(380, -120, initials +" N/A " + selected_date_string) # initial and date C - lower right box
			#can2.drawString(150, 10, initials +" N/A " + selected_date_string) # fat soluble - lower left box
			can2.restoreState()
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			outputStream2 = open(r"forms\fat_soluble\uplc_sample_form_fs_d2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			os.remove(r"forms\fat_soluble\uplc_sample_form_fs_d2.pdf")

		elif planned_testing_list[i][0] != "" and planned_testing_list[i][1] == "" and planned_testing_list[i][2] == "" and planned_testing_list[i][3] != "": # Vitamin A and D
			existing_pdf = PdfFileReader(open(r"forms\fat_soluble\uplc_sample_form_fs_ad.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)
			can2.saveState()
			can2.rotate(14)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			#This is the rotated string values for the upper portion
			can2.drawString(325, 355, initials +" N/A " + selected_date_string) # initial and date E
			can2.restoreState()

			can2.saveState()
			can2.rotate(7)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			can2.drawString(255, 185, initials +" N/A " + selected_date_string) # initial and date C - upper box
			can2.restoreState()

			can2.saveState()
			can2.rotate(30)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			can2.drawString(380, -120, initials +" N/A " + selected_date_string) # initial and date C - lower right box
			can2.restoreState()
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			outputStream2 = open(r"forms\fat_soluble\uplc_sample_form_fs_ad2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			os.remove(r"forms\fat_soluble\uplc_sample_form_fs_ad2.pdf")

		elif planned_testing_list[i][0] != "" and planned_testing_list[i][1] != "" and planned_testing_list[i][2] == "" and planned_testing_list[i][3] == "": # Vitamin A and C
			existing_pdf = PdfFileReader(open(r"forms\fat_soluble\uplc_sample_form_fs_ac.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)
			can2.saveState()
			can2.rotate(14)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			#This is the rotated string values for the upper portion
			can2.drawString(325, 355, initials +" N/A " + selected_date_string) # initial and date E
			can2.drawString(310, 230, initials +" N/A " + selected_date_string) # initial and date D
			can2.restoreState()
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			outputStream2 = open(r"forms\fat_soluble\uplc_sample_form_fs_ac2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			os.remove(r"forms\fat_soluble\uplc_sample_form_fs_ac2.pdf")

		elif planned_testing_list[i][0] != "" and planned_testing_list[i][1] == "" and planned_testing_list[i][2] != "" and planned_testing_list[i][3] == "": # Vitamin A and E
			existing_pdf = PdfFileReader(open(r"forms\fat_soluble\uplc_sample_form_fs_ae.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)
			can2.saveState()
			can2.rotate(14)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			can2.drawString(310, 230, initials +" N/A " + selected_date_string) # initial and date D
			can2.restoreState()

			can2.saveState()
			can2.rotate(7)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			can2.drawString(255, 185, initials +" N/A " + selected_date_string) # initial and date C - upper box
			can2.restoreState()

			can2.saveState()
			can2.rotate(30)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			can2.drawString(380, -120, initials +" N/A " + selected_date_string) # initial and date C - lower right box
			can2.restoreState()
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			outputStream2 = open(r"forms\fat_soluble\uplc_sample_form_fs_ae2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			os.remove(r"forms\fat_soluble\uplc_sample_form_fs_ae2.pdf")

		elif planned_testing_list[i][0] == "" and planned_testing_list[i][1] == "" and planned_testing_list[i][2] != "" and planned_testing_list[i][3] != "": # Vitamin D and E
			existing_pdf = PdfFileReader(open(r"forms\fat_soluble\uplc_sample_form_fs_de.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)
			can2.saveState()
			can2.rotate(14)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			#This is the rotated string values for the upper portion
			can2.drawString(370, 487, initials +" N/A " + selected_date_string) # initial and date A
			can2.restoreState()

			can2.saveState()
			can2.rotate(7)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			can2.drawString(255, 185, initials +" N/A " + selected_date_string) # initial and date C - upper box
			can2.restoreState()

			can2.saveState()
			can2.rotate(30)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			can2.drawString(380, -120, initials +" N/A " + selected_date_string) # initial and date C - lower right box
			can2.restoreState()
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			outputStream2 = open(r"forms\fat_soluble\uplc_sample_form_fs_de2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			os.remove(r"forms\fat_soluble\uplc_sample_form_fs_de2.pdf")

		elif planned_testing_list[i][0] == "" and planned_testing_list[i][1] != "" and planned_testing_list[i][2] != "" and planned_testing_list[i][3] == "": # Vitamin C and E
			existing_pdf = PdfFileReader(open(r"forms\fat_soluble\uplc_sample_form_fs_ce.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)
			can2.saveState()
			can2.rotate(14)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			#This is the rotated string values for the upper portion
			can2.drawString(370, 487, initials +" N/A " + selected_date_string) # initial and date A
			can2.drawString(310, 230, initials +" N/A " + selected_date_string) # initial and date D
			can2.restoreState()
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			outputStream2 = open(r"forms\fat_soluble\uplc_sample_form_fs_ce2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			os.remove(r"forms\fat_soluble\uplc_sample_form_fs_ce2.pdf")

		elif planned_testing_list[i][0] == "" and planned_testing_list[i][1] != "" and planned_testing_list[i][2] == "" and planned_testing_list[i][3] != "": # Vitamin C and D
			existing_pdf = PdfFileReader(open(r"forms\fat_soluble\uplc_sample_form_fs_cd.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)
			can2.saveState()
			can2.rotate(14)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			#This is the rotated string values for the upper portion
			can2.drawString(370, 487, initials +" N/A " + selected_date_string) # initial and date A
			can2.drawString(325, 355, initials +" N/A " + selected_date_string) # initial and date E
			can2.restoreState()
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			outputStream2 = open(r"forms\fat_soluble\uplc_sample_form_fs_cd2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			os.remove(r"forms\fat_soluble\uplc_sample_form_fs_cd2.pdf")

		elif planned_testing_list[i][0] != "" and planned_testing_list[i][1] != "" and planned_testing_list[i][2] == "" and planned_testing_list[i][3] != "": # Vitamin A, C, and D
			existing_pdf = PdfFileReader(open(r"forms\fat_soluble\uplc_sample_form_fs_acd.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)
			can2.saveState()
			can2.rotate(14)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			#This is the rotated string values for the upper portion
			can2.drawString(325, 355, initials +" N/A " + selected_date_string) # initial and date E
			can2.restoreState()
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			outputStream2 = open(r"forms\fat_soluble\uplc_sample_form_fs_acd2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			os.remove(r"forms\fat_soluble\uplc_sample_form_fs_acd2.pdf")

		elif planned_testing_list[i][0] != "" and planned_testing_list[i][1] != "" and planned_testing_list[i][2] != "" and planned_testing_list[i][3] == "": # Vitamin A, C, and E
			existing_pdf = PdfFileReader(open(r"forms\fat_soluble\uplc_sample_form_fs_ace.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)
			can2.saveState()
			can2.rotate(14)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			#This is the rotated string values for the upper portion
			can2.drawString(310, 230, initials +" N/A " + selected_date_string) # initial and date D
			can2.restoreState()
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			outputStream2 = open(r"forms\fat_soluble\uplc_sample_form_fs_ace2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			os.remove(r"forms\fat_soluble\uplc_sample_form_fs_ace2.pdf")

		elif planned_testing_list[i][0] != "" and planned_testing_list[i][1] == "" and planned_testing_list[i][2] != "" and planned_testing_list[i][3] != "": # Vitamin A, D, and E
			existing_pdf = PdfFileReader(open(r"forms\fat_soluble\uplc_sample_form_fs_ade.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)

			can2.saveState()
			can2.rotate(7)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			can2.drawString(255, 185, initials +" N/A " + selected_date_string) # initial and date C - upper box
			can2.restoreState()

			can2.saveState()
			can2.rotate(30)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			can2.drawString(380, -120, initials +" N/A " + selected_date_string) # initial and date C - lower right box
			can2.restoreState()
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			outputStream2 = open(r"forms\fat_soluble\uplc_sample_form_fs_ade2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			os.remove(r"forms\fat_soluble\uplc_sample_form_fs_ade2.pdf")

		elif planned_testing_list[i][0] == "" and planned_testing_list[i][1] != "" and planned_testing_list[i][2] != "" and planned_testing_list[i][3] != "": # Vitamin C, D, and E
			existing_pdf = PdfFileReader(open(r"forms\fat_soluble\uplc_sample_form_fs_cde.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)
			can2.saveState()
			can2.rotate(14)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			#This is the rotated string values for the upper portion
			can2.drawString(370, 487, initials +" N/A " + selected_date_string) # initial and date A
			can2.restoreState()
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			outputStream2 = open(r"forms\fat_soluble\uplc_sample_form_fs_cde2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			os.remove(r"forms\fat_soluble\uplc_sample_form_fs_cde2.pdf")

		# Water Soluble Testing Forms
		elif len(water_soluble_testing_list[i]) == 1 and triplicate_status_list[i] == True:
			existing_pdf = PdfFileReader(open(r"forms\water_soluble\uplc_sample_form_triplicate_one_ws_analyte.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)
			can2.saveState()
			can2.rotate(36)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			#This is the rotated string values for the upper portion
			can2.drawString(380, 115, initials +" N/A " + selected_date_string) # initial and date A
			can2.restoreState()
			# This is the rotated string values for the lower portion
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			outputStream2 = open(r"forms\water_soluble\uplc_sample_form_triplicate_one_ws_analyte2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			os.remove(r"forms\water_soluble\uplc_sample_form_triplicate_one_ws_analyte2.pdf")

		elif len(water_soluble_testing_list[i]) == 2 and triplicate_status_list[i] == True:
			existing_pdf = PdfFileReader(open(r"forms\water_soluble\uplc_sample_form_triplicate_two_ws_analytes.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)
			can2.saveState()
			can2.rotate(30)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			#This is the rotated string values for the upper portion
			can2.drawString(380, 125, initials +" N/A " + selected_date_string) # initial and date A
			can2.restoreState()
			# This is the rotated string values for the lower portion
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			outputStream2 = open(r"forms\water_soluble\uplc_sample_form_triplicate_two_ws_analyte2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			os.remove(r"forms\water_soluble\uplc_sample_form_triplicate_two_ws_analyte2.pdf")

		elif len(water_soluble_testing_list[i]) == 3 and triplicate_status_list[i] == True:
			existing_pdf = PdfFileReader(open(r"forms\water_soluble\uplc_sample_form_triplicate_three_ws_analytes.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)
			can2.saveState()
			can2.rotate(22)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			#This is the rotated string values for the upper portion
			can2.drawString(328, 134, initials +" N/A " + selected_date_string) # initial and date A
			can2.restoreState()
			# This is the rotated string values for the lower portion
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			outputStream2 = open(r"forms\water_soluble\uplc_sample_form_triplicate_three_ws_analyte2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			os.remove(r"forms\water_soluble\uplc_sample_form_triplicate_three_ws_analyte2.pdf")

		elif len(water_soluble_testing_list[i]) == 4 and triplicate_status_list[i] == True:
			existing_pdf = PdfFileReader(open(r"forms\water_soluble\uplc_sample_form_triplicate_four_ws_analytes.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)
			can2.saveState()
			can2.rotate(11)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			#This is the rotated string values for the upper portion
			can2.drawString(280, 155, initials +" N/A " + selected_date_string) # initial and date A
			can2.restoreState()
			# This is the rotated string values for the lower portion
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			outputStream2 = open(r"forms\water_soluble\uplc_sample_form_triplicate_four_ws_analyte2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			os.remove(r"forms\water_soluble\uplc_sample_form_triplicate_four_ws_analyte2.pdf")
		elif len(water_soluble_testing_list[i]) == 5 and triplicate_status_list[i] == True:
			existing_pdf = PdfFileReader(open(r"forms\uplc_sample_form.pdf", "rb"))
			# Do not do anything except for using the uplc form.

		elif len(water_soluble_testing_list[i]) == 1 and triplicate_status_list[i] == False:
			existing_pdf = PdfFileReader(open(r"forms\water_soluble\uplc_sample_form_single_one_ws_analyte.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)
			can2.saveState()
			can2.rotate(36)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			#This is the rotated string values for the upper portion
			can2.drawString(380, 115, initials +" N/A " + selected_date_string) # initial and date A
			can2.restoreState()
			# This is the rotated string values for the lower portion
			can2.saveState()
			can2.rotate(33)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			can2.drawString(370, -155, initials +" N/A " + selected_date_string) # initial and date for single samples - lower right box
			can2.restoreState()
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			outputStream2 = open(r"forms\water_soluble\uplc_sample_form_single_one_ws_analyte2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			os.remove(r"forms\water_soluble\uplc_sample_form_single_one_ws_analyte2.pdf")

		elif len(water_soluble_testing_list[i]) == 2 and triplicate_status_list[i] == False:
			#existing_pdf = PdfFileReader(open(r"forms\water_soluble\uplc_sample_form_single_two_ws_analytes.pdf", "rb"))
			existing_pdf = PdfFileReader(open(r"forms\water_soluble\uplc_sample_form_triplicate_two_ws_analytes.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)
			can2.saveState()
			can2.rotate(30)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			#This is the rotated string values for the upper portion
			can2.drawString(380, 125, initials +" N/A " + selected_date_string) # initial and date A
			can2.restoreState()

			# This is the rotated string values for the lower portion
			# can2.saveState()
			# can2.rotate(33)
			# can2.setFillColor(gmp_crossouts_font_color)
			# can2.setFont('Helvetica-Bold', 14)
			# can2.drawString(370, -155, initials +" N/A " + selected_date_string) # initial and date for single samples - lower right box
			# can2.restoreState()
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			#outputStream2 = open(r"forms\water_soluble\uplc_sample_form_single_two_ws_analyte2.pdf", "wb")
			outputStream2 = open(r"forms\water_soluble\uplc_sample_form_triplicate_two_ws_analyte2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			#os.remove(r"forms\water_soluble\uplc_sample_form_single_two_ws_analyte2.pdf")
			os.remove(r"forms\water_soluble\uplc_sample_form_triplicate_two_ws_analyte2.pdf")

		elif len(water_soluble_testing_list[i]) == 3 and triplicate_status_list[i] == False:
			existing_pdf = PdfFileReader(open(r"forms\water_soluble\uplc_sample_form_triplicate_three_ws_analytes.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)
			can2.saveState()
			can2.rotate(22)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			#This is the rotated string values for the upper portion
			can2.drawString(328, 134, initials +" N/A " + selected_date_string) # initial and date A
			can2.restoreState()
			# This is the rotated string values for the lower portion
			# can2.saveState()
			# can2.rotate(33)
			# can2.setFillColor(gmp_crossouts_font_color)
			# can2.setFont('Helvetica-Bold', 14)
			# can2.drawString(370, -155, initials +" N/A " + selected_date_string) # initial and date for single samples - lower right box
			# can2.restoreState()
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			outputStream2 = open(r"forms\water_soluble\uplc_sample_form_triplicate_three_ws_analyte2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			os.remove(r"forms\water_soluble\uplc_sample_form_triplicate_three_ws_analyte2.pdf")

		elif len(water_soluble_testing_list[i]) == 4 and triplicate_status_list[i] == False:
			existing_pdf = PdfFileReader(open(r"forms\water_soluble\uplc_sample_form_triplicate_four_ws_analytes.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)
			can2.saveState()
			can2.rotate(11)
			can2.setFillColor(gmp_crossouts_font_color)
			can2.setFont('Helvetica-Bold', 14)
			#This is the rotated string values for the upper portion
			can2.drawString(280, 155, initials +" N/A " + selected_date_string) # initial and date A
			can2.restoreState()
			# This is the rotated string values for the lower portion
			# can2.saveState()
			# can2.rotate(33)
			# can2.setFillColor(gmp_crossouts_font_color)
			# can2.setFont('Helvetica-Bold', 14)
			# can2.drawString(370, -155, initials +" N/A " + selected_date_string) # initial and date for single samples - lower right box
			# can2.restoreState()
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			outputStream2 = open(r"forms\water_soluble\uplc_sample_form_triplicate_four_ws_analyte2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			os.remove(r"forms\water_soluble\uplc_sample_form_triplicate_four_ws_analyte2.pdf")

		elif len(water_soluble_testing_list[i]) == 5 and triplicate_status_list[i] == False:
			existing_pdf = PdfFileReader(open(r"forms\water_soluble\uplc_sample_form_triplicate_five_ws_analytes.pdf", "rb"))
			packet2 = io.BytesIO()
			can2 = canvas.Canvas(packet2, pagesize=letter)
			packet2.seek(0)
			# This is the rotated string values for the lower portion
			# can2.saveState()
			# can2.rotate(33)
			# can2.setFillColor(gmp_crossouts_font_color)
			# can2.setFont('Helvetica-Bold', 14)
			# can2.drawString(370, -155, initials +" N/A " + selected_date_string) # initial and date for single samples - lower right box
			# can2.restoreState()
			can2.save()
			packet2.seek(0)
			new_pdf2 = PdfFileReader(packet2)
			output2 = PdfFileWriter()
			if polarities[i] == "Fat Soluble":
				page_A = existing_pdf.getPage(0)
			else:
				page_A = existing_pdf.getPage(1)
			page_B = new_pdf2.getPage(0)
			page_A.mergePage(page_B)
			output2.addPage(page_A)
			outputStream2 = open(r"forms\water_soluble\uplc_sample_form_triplicate_five_ws_analyte2.pdf", "wb")
			output2.write(outputStream2)
			outputStream2.close()
			os.remove(r"forms\water_soluble\uplc_sample_form_triplicate_five_ws_analyte2.pdf")

		else:
			existing_pdf = PdfFileReader(open(r"forms\uplc_sample_form.pdf", "rb"))

		output = PdfFileWriter()
		if polarities[i] == "Fat Soluble":
			page = existing_pdf.getPage(0)
		else:
			page = existing_pdf.getPage(1)
		page2 = new_pdf.getPage(0)
		page.mergePage(page2)
		output.addPage(page)
		new_pdf_name = wips[i] + "_sample_form" + "_(" + str(wips[i]) + ")" + ".pdf"
		outputStream = open(new_pdf_name, "wb")
		output.write(outputStream)
		outputStream.close()

		merger.append(PdfFileReader(new_pdf_name, 'rb'))
		#deletes initial merged pdf after appending to the batch pdf
		os.remove(new_pdf_name)
		i += 1
	# Merge all of the single PDF documents together into a single pdf document for easy printing
	merger.write(batch_file_name)

	# logic controlling how files are organized in the sub-directory 'uplc batches' and makes sure more than one batch with the same name can be created and opened.
	sub_dir = r"\uplc batches\\"
	# if os.path.exists(os.path.join(os.getcwd(), sub_dir)) == False:
	# 	os.system('mkdir ./uplc_batches')
	if path.exists(os.getcwd() + sub_dir + batch_file_name) == False:
		# move batch file to the uplc batches directory
		os.rename(batch_file_name, os.getcwd() + sub_dir + batch_file_name)
		# open the batch file
		os.startfile(os.getcwd() + sub_dir + batch_file_name) # The os.startfile will only work on Windows Machines # This line will open newly created PDF file
	else:
		file_count = 0
		for files in os.walk('./uplc batches/'):
			pattern = re.compile(batch_file_name_no_pdf)
			for each in files[2]:
				if pattern.search(each) != None:
					file_count += 1
				else:
					file_count += 0
		file_num = str(file_count + 1)
		os.rename(batch_file_name, os.getcwd() + sub_dir + batch_file_name_no_pdf + "_" + file_num + ".pdf")
		os.startfile(os.getcwd() + sub_dir + batch_file_name_no_pdf + "_" + file_num + ".pdf")

	print("\n********************************************************************done.\n")

def clear():
	#clear the wip entry boxes
	wip_list = [wip1, wip2, wip3, wip4, wip5, wip6, wip7]
	for wip in wip_list:
		wip.delete(0, END)

	#clear the lot extry boxes
	lot_list = [lot1, lot2, lot3, lot4, lot5, lot6, lot7]
	for lot in lot_list:
		lot.delete(0, END)

	#clear the description boxes
	description_list = [desc1, desc2, desc3, desc4, desc5, desc6, desc7]
	for desc in description_list:
		desc.delete(0, END)

	#clear the lab reports boxes
	lab_reports_list = [lr1, lr2, lr3, lr4, lr5, lr6 ,lr7]
	for lab_report in lab_reports_list:
		lab_report.delete(0, END)

	#clear the sample type dropdowns
	sample_type_list = [st1, st2, st3, st4, st5, st6, st7]
	for sample_type in sample_type_list:
		sample_type['state'] = 'normal'
		sample_type.delete(0, END)
		sample_type['state'] = 'readonly'

	#clear the sample polarity dropdowns
	polarity_list = [polarity1, polarity2, polarity3, polarity4, polarity5, polarity6, polarity7]
	for polarity in polarity_list:
		polarity['state'] = 'normal'
		polarity.delete(0, END)
		polarity['state'] = 'readonly'

	#clear the sample weight fields
	weight_list = [wt1, wt2, wt3, wt4, wt5, wt6, wt7]
	for wt in weight_list:
		wt.delete(0, END)

	#clear the units per serving fields
	ups_list = [ups1, ups2, ups3, ups4, ups5, ups6, ups7]
	for ups in ups_list:
		ups.delete(0, END)

	#set the triplicate checkboxes back to unchecked
	triplicate_checkbox_list = [chk_tri1, chk_tri2, chk_tri3, chk_tri4, chk_tri5, chk_tri6, chk_tri7]
	for triplicate_checkbox in triplicate_checkbox_list:
		triplicate_checkbox.set(False)

	# Set the stability checkboxes back to unchecked
	stability_checkbox_list = [stability_box1, stability_box2, stability_box3, stability_box4, stability_box5, stability_box6, stability_box7]
	for stability_checkbox in stability_checkbox_list:
		stability_checkbox.set(False)

	#clear the first set of vitamin type dropdowns
	r1_fs1.delete(0, END)
	r1_fs2.delete(0, END)
	r1_fs3.delete(0, END)
	r1_fs4.delete(0, END)
	r1_ws1.delete(0, END)
	r1_ws2.delete(0, END)
	r1_ws3.delete(0, END)
	r1_ws4.delete(0, END)
	r1_ws5.delete(0, END)
	#clear the second set of vitamin type dropdowns
	r2_fs1.delete(0, END)
	r2_fs2.delete(0, END)
	r2_fs3.delete(0, END)
	r2_fs4.delete(0, END)
	r2_ws1.delete(0, END)
	r2_ws2.delete(0, END)
	r2_ws3.delete(0, END)
	r2_ws4.delete(0, END)
	r2_ws5.delete(0, END)
	#clear the third set of vitamin type dropdowns
	r3_fs1.delete(0, END)
	r3_fs2.delete(0, END)
	r3_fs3.delete(0, END)
	r3_fs4.delete(0, END)
	r3_ws1.delete(0, END)
	r3_ws2.delete(0, END)
	r3_ws3.delete(0, END)
	r3_ws4.delete(0, END)
	r3_ws5.delete(0, END)
	#clear the fourth set of vitamin type dropdowns
	r4_fs1.delete(0, END)
	r4_fs2.delete(0, END)
	r4_fs3.delete(0, END)
	r4_fs4.delete(0, END)
	r4_ws1.delete(0, END)
	r4_ws2.delete(0, END)
	r4_ws3.delete(0, END)
	r4_ws4.delete(0, END)
	r4_ws5.delete(0, END)
	#clear the fifth set of vitamin type dropdowns
	r5_fs1.delete(0, END)
	r5_fs2.delete(0, END)
	r5_fs3.delete(0, END)
	r5_fs4.delete(0, END)
	r5_ws1.delete(0, END)
	r5_ws2.delete(0, END)
	r5_ws3.delete(0, END)
	r5_ws4.delete(0, END)
	r5_ws5.delete(0, END)
	#clear the sixth set of vitamin type dropdowns
	r6_fs1.delete(0, END)
	r6_fs2.delete(0, END)
	r6_fs3.delete(0, END)
	r6_fs4.delete(0, END)
	r6_ws1.delete(0, END)
	r6_ws2.delete(0, END)
	r6_ws3.delete(0, END)
	r6_ws4.delete(0, END)
	r6_ws5.delete(0, END)
	#clear the seventh set of vitamin type dropdowns
	r7_fs1.delete(0, END)
	r7_fs2.delete(0, END)
	r7_fs3.delete(0, END)
	r7_fs4.delete(0, END)
	r7_ws1.delete(0, END)
	r7_ws2.delete(0, END)
	r7_ws3.delete(0, END)
	r7_ws4.delete(0, END)
	r7_ws5.delete(0, END)

def openOldPdfFileMenu():
	form_button_frame.filename = filedialog.askopenfilename(initialdir=r"./uplc batches", title="UPLC Batch PDF Menu", filetypes=(("PDF files", "*.pdf"), ("All files", "*.*")))
	selected_file = form_button_frame.filename
	os.startfile(selected_file)

def check_specifications():

	print(f"{Fore.RED}   _____  _____ _   _    _____                  _____ _               _             ")
	print(f"{Fore.YELLOW}  / ____|/ ____| \ | |  / ____|                / ____| |             | |            ")
	print(f"{Fore.GREEN} | (___ | |    |  \| | | (___  _ __   ___  ___| |    | |__   ___  ___| | _____ _ __ ")
	print(f"{Fore.BLUE}  \___ \| |    | . ` |  \___ \| '_ \ / _ \/ __| |    | '_ \ / _ \/ __| |/ / _ \ '__|")
	print(f"{Fore.MAGENTA}  ____) | |____| |\  |  ____) | |_) |  __/ (__| |____| | | |  __/ (__|   <  __/ |   ")
	print(f"{Fore.RED} |_____/ \_____|_| \_| |_____/| .__/ \___|\___|\_____|_| |_|\___|\___|_|\_\___|_|   ")
	print(f"{Fore.YELLOW}                              | |                                                   ")
	print(f"{Fore.GREEN}                              |_|                                                   ")
	print(f"{Fore.RESET}                                                                                    ")

	try:
		# Read in DataFrame and make sure dtype is string.
		#df = pd.read_excel(os.path.normpath('R:\\QC\\Laboratory\\7. Lab improvement\\Raw Material Testing Plan\\test\\specification_database.xlsx'), engine='openpyxl', dtype=str)
		df = read_encrypted_excel(os.path.join(R_drive_parent_folder,'specification_database.xlsx'))
		print(f"Reading specification_database.xlsx from {Fore.GREEN}{R_drive_parent_folder}{Fore.RESET}.")
	except:
		# Read in DataFrame and make sure dtype is string.
		print(f"{Fore.YELLOW}Could not access 'specification_database.xlsx' in the R: drive or password is incorrect.{Fore.RESET} Trying local location now.")
		try:
			df = pd.read_excel(os.path.join(os.getcwd(), 'specification_database.xlsx'), engine='openpyxl', dtype=str)
			print(f"'specification_database.xlsx' {Fore.GREEN}successfully found on local drive.{Fore.RESET}")
		except:
			print(f"\n\n{Fore.RED}***************************************************************************************{Fore.RESET}")
			print(f"{Fore.RED}CRITICAL WARNING!:{Fore.RESET}")
			print(f"{Fore.RED}Failed to find 'specification_database.xlsx' in the R drive or on your local computer.{Fore.RESET}")
			print(f"{Fore.RED}***************************************************************************************{Fore.RESET}\n\n")


	def return_spec_value(wip, spec_database_property_column):
		spec_database_property_column = str(spec_database_property_column)
		# Read spec database in R drive if available else local version
		try:
			# Read in DataFrame and make sure dtype is string.
			#df = read_encrypted_excel(os.path.normpath('R:\\QC\\Laboratory\\7. Lab improvement\\Raw Material Testing Plan\\test\\specification_database.xlsx'))
			df = read_encrypted_excel(os.path.join(R_drive_parent_folder, 'specification_database.xlsx'))
			#df = pd.read_excel(os.path.join(R_drive_parent_folder, 'specification_database.xlsx'), engine='openpyxl', dtype=str)
		except:
			# Read in DataFrame and make sure dtype is string.
			print(f"{Fore.RED}Could not access 'specification_database.xlsx' in the R: drive.{Fore.RESET} Trying local location now.")
			df = pd.read_excel(os.path.join(os.getcwd(), 'specification_database.xlsx'), engine='openpyxl', dtype=str)

		# Convert all values contained within the specification excel to string type variables
		# Make a new data frame that only contains WIP equal to the entered WIP value in the GUI
		df2 = df.loc[df['WIP/Item'] == str(wip)]

		# Reset the index values so that we can reliably pick the 0th row (first raw values)
		df2.reset_index(inplace=True)

		# If a matching WIP was found
		if df2.empty == False:
			return(df2.iloc[0][spec_database_property_column])
		else:
			return('')

	wip_fields = [wip1, wip2, wip3, wip4, wip5, wip6, wip7]

	get_small_focus_list = [
		['.!frame.!entry3', '.!frame.!combobox', '.!frame.!entry5', '.!frame.!entry6'],
		['.!frame.!entry9', '.!frame.!combobox3', '.!frame.!entry11', '.!frame.!entry12'],
		['.!frame.!entry15', '.!frame.!combobox5', '.!frame.!entry17', '.!frame.!entry18'],
		['.!frame.!entry21', '.!frame.!combobox7', '.!frame.!entry23', '.!frame.!entry24'],
		['.!frame.!entry27', '.!frame.!combobox9', '.!frame.!entry29', '.!frame.!entry30'],
		['.!frame.!entry33', '.!frame.!combobox11', '.!frame.!entry35', '.!frame.!entry36'],
		['.!frame.!entry39', '.!frame.!combobox13', '.!frame.!entry41', '.!frame.!entry42']
	]

	set_small_focus_list = [
		[desc1, st1, wt1, ups1],
		[desc2, st2, wt2, ups2],
		[desc3, st3, wt3, ups3],
		[desc4, st4, wt4, ups4],
		[desc5, st5, wt5, ups5],
		[desc6, st6, wt6, ups6],
		[desc7, st7, wt7, ups7]
	]

	for row, wip_field in enumerate(wip_fields):
		# If a WIP has been entered in the GUI row
		if str(wip_field.get()) != '':
			wip = str(wip_field.get())
			print("\n****************************************************")

			for col, field in enumerate(set_small_focus_list[row]):
				# If field is empty check to see if spec is available
				# Description Fields
				if col == 0:
					# If description field has something written in it set description variable to the text that is entered there
					if field.get() != '':
						description = field.get()

					# If description field has nothing written in it, find description in database and set description variable to that value.
					if field.get() == '':
						description = return_spec_value(wip, 'Description')
						if description != '':
							field.insert(0, description)

				# Sample Type Fields
				if col == 1:
					# Retreive Sample Type from database
					sample_type = return_spec_value(wip, 'Sample Type')

					if sample_type != '':
						if sample_type == 'Finished Product':
							field.current(1)
						if sample_type == 'Raw Blend':
							field.current(2)
						if sample_type == 'Percent Active':
							field.current(3)

					if sample_type == '':
						if field.get() == '':
							print(f'\n{Fore.RED}Specification database does not have a sample type defined for WIP {Fore.YELLOW}{wip}{Fore.RESET}.{Fore.RED}\nThis is a strong indication that the WIP is not entered into the specification database{Fore.RESET}.\nIf you think this is an error, please remember to manually set the sample type in the form before making your batch.\n')

				# Piece Weight fields
				if col == 2:
					piece_weight = return_spec_value(wip, 'Piece Weight/Usage')
					try:
						if piece_weight != '':
							# if the field contains information, delete the info and update with info from database.
							if field.get() != '':
								field.delete(0, END)
								field.insert(0, piece_weight)
							else:
								if pd.isnull(piece_weight):
									field.insert(0, '')
								else:
									field.insert(0, piece_weight)
					except:
							print(f"{Fore.RED}Please remember to enter the Piece Weight for WIP {wip}{Fore.RESET}")
							pass

				if col == 3:
					units_per_spec = return_spec_value(wip, 'spec serving')
					try:
						if units_per_spec != '':
							# if the field contains information, delete the info and update with info from database.
							if field.get() != '':
								field.delete(0, END)
								field.insert(0, units_per_spec)
							else:
								if pd.isnull(units_per_spec):
									field.insert(0, '')
								else:
									field.insert(0, units_per_spec)
					except:
						print(f"{Fore.RED}Usage per spec data is missing from the specification database.\nPlease manually enter the 'Usage Per Spec' information.{Fore.RESET}")
						pass

				piece_weight_units = return_spec_value(wip, 'PW/Usage Units')

			print(f"WIP/Item: {wip}")
			if description == '' or pd.isnull(description):
				print(f"Sample Description: {Fore.RED}MISSING{Fore.RESET}")
			else:
				print(f"Sample Description: {description}")
			if sample_type == '' or pd.isnull(sample_type):
				print(f"Sample Type: {Fore.RED}MISSING{Fore.RESET}")
			else:
				print(f"Sample Type: {sample_type}")
			if piece_weight == '' or pd.isnull(piece_weight):
				print(f"Piece Weight / Usage: {Fore.RED}MISSING{Fore.RESET}")
			else:
				print(f"Piece Weight / Usage: {piece_weight} {piece_weight_units}")
			if units_per_spec == '' or pd.isnull(units_per_spec):
				print(f"Units Per Spec: {Fore.RED}MISSING{Fore.RESET}")
			else:
				print(f"Units Per Spec: {units_per_spec}")


	print("\n***********************************************done.")

form_button_frame = Frame(root, width=100, height=200)
form_button_frame.grid(column=1, row=2)

populate_specs = Button(form_button_frame, text="Check Specs", command=check_specifications)
populate_specs.grid(column=0, row=0, padx=2, pady=20)

submit_btn = Button(form_button_frame, text="Make PDF", command=click_and_exit)
submit_btn.grid(column=1, row=0, padx=2, pady=20)

clear_btn = Button(form_button_frame, text="Clear Fields", command=clear)
clear_btn.grid(column=2, row =0, padx=2, pady=20)

# old_pdf_btn = Button(form_button_frame, text="Find Batch", command=openOldPdfFileMenu)
# old_pdf_btn.grid(column=2, row=0, padx=2, pady=20)

root.mainloop()
